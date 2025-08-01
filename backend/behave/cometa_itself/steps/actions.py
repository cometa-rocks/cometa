# -*- coding: UTF-8 -*-
# -------------------------------------------------------
# This archive contains all the steps available in Cometa Front for execution.
# Steps not included are Enterprise Licensed Steps
# Any change in this file are automatically reflected in Cometa on next execution of a feature.
#
# Changelog
# 2024-03-01 RRO Added delimiter sniffing to reading CSV files
# 2022-07-11 RRO Added some sleeps of 100ms to copy and saving of downloaded and edited excel files, as received IO timeouts
# 2022-07-08 RRO added last_downloaded_file.suffix to handle generated generic filenames where the suffix is maintained
# 2022-03-04 RRO added new step "Search for "{something}" in IBM Cognos and click on first result"
# 2022-03-01 RRO added step to hit ok on alert, confirm or prompt window
# 2022-02-01 RRO Cleaning up :-)
# 2017-11-21 Compare.sh fuer besseren  Vergleich von Images angelegt
# 2017-11-12 PoC Arbeiten abgeschlossen
#
# -------------------------------------------------------
from behave import *
import time
import sys
import os, glob
import os.path
import shutil
import subprocess
import shlex
import requests
import json
import re
import csv
import datetime
import signal
import logging
import traceback
import urllib.parse
import random
import copy
# import PIL
from subprocess import call, run
from selenium.common.exceptions import WebDriverException, NoAlertPresentException, ElementNotInteractableException, TimeoutException, StaleElementReferenceException
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.keys import Keys
from functools import wraps
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.actions.wheel_input import ScrollOrigin
from selenium.webdriver.remote.file_detector import LocalFileDetector
# Import utilities
sys.path.append(os.path.dirname(os.path.realpath(__file__)))
from tools.cognos import *
from tools.common import *
from tools.exceptions import *
from tools.variables import *
from slugify import slugify
from html_diff import diff
from bs4 import BeautifulSoup

# pycrypto imports
from Crypto import Random
from Crypto.Cipher import AES
import base64
base64.encodestring = base64.encodebytes
from hashlib import md5
from pathlib import Path
from tools import expected_conditions as CEC
from utility.config_handler import *
import sys

sys.path.append("/opt/code/behave_django")
sys.path.append('/opt/code/cometa_itself/steps')

from utility.functions import *
from utility.configurations import ConfigurationManager
from utility.common import *
from utility.encryption import *
from tools.common_functions import *

SCREENSHOT_PREFIX = ConfigurationManager.get_configuration('COMETA_SCREENSHOT_PREFIX', '')
ENCRYPTION_START = ConfigurationManager.get_configuration('COMETA_ENCRYPTION_START', '')

# setup logging
logger = logging.getLogger('FeatureExecution')




# Browse to an URL
# Example: StartBrowser and call URL "https://prod.cometa.rocks/#/new"
@step(u'StartBrowser and call URL "{url}"')
@done(u'StartBrowser and call URL "{url}"')
def step_impl(context,url):
    send_step_details(context, 'Loading page')
    context.browser.get(url)

# Browse to an URL
# Example: StartBrowser and call URL "https://prod.cometa.rocks/#/new"
@step(u'StartBrowser and call URL "{url}" with javascript')
@done(u'StartBrowser and call URL "{url}" with javascript')
def step_impl(context,url):
    send_step_details(context, 'Loading page')
    context.browser.execute_script(f"window.location.href = '{url}'")

# Browse to an URL
# Example: Goto URL "https://prod.cometa.rocks/#/new"
@step(u'Goto URL "{url}"')
@done(u'Goto URL "{url}"')
def step_impl(context,url):
    send_step_details(context, 'Loading page')
    context.browser.get(url)

# Sets a comma separated list of environments
# Example: a set of environments
@step(u'a set of environments')
@done(u'a set of environments')
def step_impl(context):
    context.environments = []
    for row in context.table:
        this_env = [row[0],row]
        logger.debug("row:",row)
        logger.debug("env:",this_env)
        context.environments.append(this_env)
    if context.table.count() == 0:
        raise CustomError("No environments found.")

# Clears all cookies from the browser session
# Example: This step deletes all cookies for the current browser session
@step(u'Delete all cookies')
@done(u'Delete all cookies')
def step_impl(context):
    print("Delete all cookies")
    send_command = ('POST', '/session/$sessionId/chromium/send_command')
    context.browser.command_executor._commands['SEND_COMMAND'] = send_command
    context.browser.execute('SEND_COMMAND', dict(cmd='Network.clearBrowserCookies', params={}))

# Moves the mouse to the css selector and clicks
# Example: I move mouse to "//div[contains(@routerlink, '/')]" and click
@step(u'I move mouse to "{css_selector}" and click')
@done(u'I move mouse to "{css_selector}" and click')
def step_impl(context,css_selector):
    send_step_details(context, 'Looking for selector')
    elem = waitSelector(context, "css", css_selector)
    generate_dataset = context.feature_info.get('generate_dataset', False)
    logger.debug(f"Generate Dataset: {str(generate_dataset)}")
    payload = {
        'context': base64.b64encode(clear_html_page_source(context.browser.page_source).encode()).decode(),
        'Target_Source': base64.b64encode(clear_html_page_source(elem[0].get_attribute('outerHTML')).encode()).decode(),
        'target': css_selector,
        'success': False,
        'feature_result_id': os.environ['feature_result_id']
    }
    send_step_details(context, 'Clicking')
    try:
        ActionChains(context.browser).move_to_element(elem[0]).click().perform()
        if generate_dataset:
            payload['success'] = True
            requests.post(f'{get_cometa_backend_url()}/api/dataset/', headers={"Host": "cometa.local"}, json=payload)
    except Exception as err:
        if isCommandNotSupported(err):
            # I move mouse is not supported in the current device, falling back to "click element with css"
            send_step_details(context, 'Incompatible step, falling back to "I click on css selector"')
            click_element_by_css(context, css_selector)

            if generate_dataset:
                payload['success'] = False
                requests.post(f'{get_cometa_backend_url()}/api/dataset/', headers={"Host": "cometa.local"}, json=payload)
        else:
            raise err

# Moves the mouse to the css selector and double clicks
# Example: I move mouse to "//div[contains(@class, 'showFolders')]" and double click
@step(u'I move mouse to "{selector}" and double click')
@done(u'I move mouse to "{selector}" and double click')
def step_impl(context,selector):
    send_step_details(context, 'Looking for selector')
    elem = waitSelector(context, "css", selector)
    send_step_details(context, 'Double Clicking')
    try:
        ActionChains(context.browser).move_to_element(elem[0]).double_click().perform()
    except Exception as err:
        logger.error("Unable to double click on the element.")
        logger.exception(err)
        raise err

# Moves the mouse to the center of css selector
# Example: I move mouse over "(//mat-icon[text()='home'])[1]"
@step(u'I move mouse over "{css_selector}"')
@done(u'I move mouse over "{css_selector}"')
def step_impl(context,css_selector):
    send_step_details(context, 'Looking for selector')
    elem = waitSelector(context, "css", css_selector)
    send_step_details(context, 'Moving Mouse')
    ActionChains(context.browser).move_to_element(elem[0]).perform()

# Moves the mouse to the center of css selector
# Example: I move mouse to "(//mat-icon[text()='home'])[1]" and right-click 
@step(u'I move mouse to "{css_selector}" and right-click')
@done(u'I move mouse to "{css_selector}" and right-click')
def step_impl(context, css_selector):
    send_step_details(context, 'Looking for selector')
    elem = waitSelector(context, "css", css_selector)
    send_step_details(context, 'Right Clicking')
    ActionChains(context.browser).context_click(elem[0]).perform()

# Moves the mouse to random element in selector and click
# Example: I can move mouse and click randomly "5" times on elements in "//cometa-folder-item-tree")
@step(u'I can move mouse and click randomly "{x}" times on elements in "{selector}"')
@done(u'I can move mouse and click randomly "{x}" times on elements in "{selector}"')
def step_impl(context, x, selector):
    send_step_details(context, 'Looking for selector')
    elements = waitSelector(context, "css", selector)

    for i in range(int(x)):
        # get a random number between 0 and elements
        index = random.randint(0, (len(elements) - 1))
        # get element
        element = elements[index]
        send_step_details(context, 'Randomly clicking on element at index %d' % index)
        try:
            # display element to the screen
            context.browser.execute_script("arguments[0].scrollIntoView({behavior: 'auto',block: 'center',inline: 'center'});", element)
            ActionChains(context.browser).move_to_element(element).click().perform()
        except Exception as err:
            if isCommandNotSupported(err):
                # I move mouse is not supported in the current device, falling back to "click element with css"
                send_step_details(context, 'Incompatible step, falling back to "I click on css selector"')
                click_element(context, element)
            else:
                raise err
        # let the page breathe a bit
        time.sleep(0.5)

# Set Environment ID
# Example: Environment "Default"
@step('Environment "{env}"')
@done('Environment "{env}"')
def step_impl(context,env):
    for item in context.environments:
        if (item[0] == env):
            context.active_environment = item[1]

# Search for something in IBM Cognos and click on the first result. Tested & Works with CA 11.1 & 11.2
# Example: Search for "Quarterly Sales Report" in IBM Cognos and click on first result
@step(u'Search for "{something}" in IBM Cognos and click on first result')
@done(u'Search for "{something}" in IBM Cognos and click on first result')
def step_impl(context, something):
    logger.debug("Searching for %s in IBM Cognos" % something )
    send_step_details(context, 'Searching for %s in IBM Cognos' % something )

    # first try to find and click a CA11.1 or CA11.2 searchbox
    # ... then select everything with CTRL+A and send the new searchstring
    # ... Hit Enter and click on the first item in the search result

    # Looking for the searchbox
    try:
        # CA11.1 search box
        logger.debug("Trying to search on CA11.1")
        elm = context.browser.find_element(By.XPATH, "//*[@id='com.ibm.bi.search.search']")
        logger.debug("elm returned")
        logger.debug("Got elm %s " % elm )
        elm.click()
        # select the opening slide input search element
        elm = waitSelector(context, "//input[@type='search']")[0]
        elm.click()
    except:
        try:
            # CA11.2 search box
            logger.debug("Trying to search on CA11.2")

            # Click on search icon - just in case browser is set to small to click on the input
            try:
                context.browser.find_element(By.XPATH, "//div[@role='search']").click()
                logger.debug("Clicked search Icon ... searchbox should now be visible.")
            except:
                logger.debug("Tried to click search icon ... but failed. This happens, when browser is maximize and searchbox is visible.")

            # wait for animation to stop
            time.sleep(0.01)

            # now click in input
            elms = context.browser.find_elements(By.XPATH, "//input[@role='searchbox']")
            logger.debug("Found %s Searchboxes" % len(elms))
            # we might need to loop here as Cognos produces various searchboxes the HTML
            for elm in elms:
                try:
                    elm.click()
                    logger.debug("==> Clicked on Element %s %s " % ( elm, elm.get_attribute("outerHTML") ) )
                except:
                    logger.debug("==> Element %s %s not clickable" % ( elm, elm.get_attribute("outerHTML") ) )
        except:
            logger.debug("Could not find a CA11.1 or CA11.2 compatible searchbox.")
            raise CustomError("Could not find a CA11.1 or CA11.2 compatible searchbox. Please have a look at the logfiles.")

    # sleep 10ms for cognos to react on the last click
    time.sleep(0.01)

    # Press Control+A to mark whatever is in the searchbox
    elm.send_keys(Keys.CONTROL, "a")
    # send the search text
    elm.send_keys(something)
    elm.send_keys(Keys.ENTER)

    # wait half a second for results to appear
    time.sleep(0.5)

    # get the first result and click on it
    logger.debug("Trying to clicking on first result")
    try:
        waitSelector(context, "xpath", "//td//div[contains(.,'%s')]" % something)[0].click()
    except:
        logger.debug("Could not click on search result - maybe there was nothing found?")
        raise CustomError("Could not click on search result - maybe there was nothing found? Please check the video or screenshots.")

    # wait 250ms a second for results to appear
    time.sleep(0.25)

# Allows to navigate to a folder in an IBM Cognos installation
# Example: I can go to IBM Cognos folder "IBM Tech"
@step(u'I can go to IBM Cognos folder "{folder_name}"')
@done(u'I can go to IBM Cognos folder "{folder_name}"')
def step_impl(context, folder_name):
    logger.debug("Splitting folders by ';' -  %s " % folder_name)
    # split folder_name by ";" to see if user wants to go into sub-folders
    folders = folder_name.split(";")
    # Clear filters if any
    send_step_details(context, 'Clearing any folder filters')
    logger.debug("Clearing folder filters if any before opening folder.")
    clearFolderFilters(context)
    # loop over each folder in order to get to the sub-folder requested by user
    for f in folders:
        logger.debug("Searching for folder: %s" % f)
        send_step_details(context, 'Searching folder')
        # Select folder
        folder = search_folder(context, f)
        logger.debug("Clicking on folder: %s" % f)
        # Open selected folder
        folder.click()
        # Allow 1 second for changing between folders
        time.sleep(0.5)
        logger.debug("Cleaning filters.")
        # Clear search filter
        clearFolderFilters(context)
        # sleep for 1s
        time.sleep(0.5)

# waitFor(element) .... waits until innerText attribute of element contains something
# @timeout("Couldn't find any text inside report for <seconds> seconds.")
def wait_for_element_text(context, css_selector):
    while True:
        elements = context.browser.find_elements(By.CSS_SELECTOR, css_selector)
        if len(elements) > 0:
            inner_text = elements[0].get_attribute('innerText')
            if inner_text:
                return True
        time.sleep(1)
    return False

def loopSubFolders(context):
    # wait until the current folder is loaded
    # wait until the current folder is loaded
    waitSelector(context, "xpath", '//*[@class="dataTables_scrollBody"]//tbody[count(tr) > 0]|//*[@class="dataTables_empty"]')
    # Make sure all folder items are in the list
    cognos_scroll_folder_till_bottom(context, True)
    # get all sub-folders inside current folder
    subFolders = context.browser.find_elements(By.XPATH, '//*[@title="Folder" or @title="Package"]/ancestor::tr//*[contains(@class, "name")]/div')
    # get all reports in current folder
    reports = context.browser.find_elements(By.XPATH, '//*[@title="Report" or @title="Query"]/ancestor::tr//*[contains(@class, "name")]/div')
    # loop over each folder and find subFolders
    for index, sF in enumerate(subFolders):
        # get the subFolder again using index just in case cognos was reloaded and we lost the refrence
        subFolder = waitSelector(context, "xpath", '//*[@title="Folder" or @title="Package"]/ancestor::tr//*[contains(@class, "name")]/div')
        subFolder = subFolder[index]
        subFolderName = subFolder.get_attribute("innerText")
        # enter subFolder
        subFolder.click()
        # self call to keep going deep
        repCount = loopSubFolders(context)
        print("Found %d reports in Folder %s" % ( repCount, subFolderName))
        # done looping subFolder
        # go back to parent folder
        previousButton = waitSelector(context, "css", '[aria-label="Previous"]')
        previousButton[0].click()
        waitSelector(context, "xpath", '//*[@class="dataTables_scrollBody"]//tbody[count(tr) > 0]|//*[@class="dataTables_empty"]')
        # Make sure all folder items are in the list
        cognos_scroll_folder_till_bottom(context, True)

    return len(reports)

# Testing with folder structure
# Example: How many reports are in current IBM Cognos Folder
@step(u'How many reports are in current IBM Cognos Folder')
@done(u'How many reports are in current IBM Cognos Folder')
def step_imp(context):
    open_team_folder(context)
    cognos_scroll_folder_till_bottom(context, True)
    loopSubFolders(context)

# Allows to test all reports inside a folder of a Cognos installation
# Example: I can test current IBM Cognos folder
@step(u'I can test current IBM Cognos folder')
@done(u'I can test current IBM Cognos folder')
def step_impl(context):
    test_folder_aso(context)

# Allows to click on the OK button of an alert, confirm or prompt message
# Example: I can click OK on alert, confirm or prompt message
@step(u'I can click OK on alert, confirm or prompt message')
@done(u'I can click OK on alert, confirm or prompt message')
def step_impl(context):
    try:
        # try switching to the alert windows
        alert = context.browser.switch_to.alert
        # then hit the enter key
        alert.accept()
    except NoAlertPresentException:
        logger.debug("Could not find the alert, confirm or prompt window")
        raise CustomError('There was no alert, confirm or prompt window present.')

#
# FIXME documentation
#
def test_folder(context, parameters = {}):

    # Make sure Team Folder is opened
    open_team_folder(context)
    send_step_details(context, 'Opening team folder')

    # Make sure all folder items are in the list
    send_step_details(context, 'Listing all folder items')
    cognos_scroll_folder_till_bottom(context, True)

    # Get all reports in current folder
    elements = waitSelector(context, "xpath", '//div[@id="teamFoldersSlideoutContent"]/descendant::div[@title and @class="nameColumnDiv contentListFocusable clickable active"]')

    # Get only 5 first elements for testing
    # elements = elements[:5]
    total = len(elements)
    if len(elements) == 0:
        raise CustomError('Empty folder')

    # Test every report by clicking on it, run it, check content, and finally close
    logger.debug("Looping over folder content with %d elements" % len(elements))

    # Set the over_results_ok_variable to true
    # ... if one report fails, then this will change to false
    over_all_results_ok=True

    # Loop over elements in folder
    for index, el in enumerate(elements):
        print(' ==> Testing report #%s' % str(index))
        logger.debug(' ==================')
        logger.debug(' ==> Testing report #%s of %s' % (str(index),str(total)))
        logger.debug(' ==================')
        report_start_time=time.time()
        # Make sure Team Folder is opened
        logger.debug("makeing sure team folder is opened")
        open_team_folder(context)
        # Make sure all folder items are in the list
        logger.debug("Scrolling folder to bottom")
        cognos_scroll_folder_till_bottom(context, True)
        # Retrieve element again as DOM can later be changed
        logger.debug("Getting DOM element %d in folder" % (int(index)+1))
        element = context.browser.find_element(By.XPATH, '//div[@id="teamFoldersSlideoutContent"]/descendant::div[@title and @class="nameColumnDiv contentListFocusable clickable active"][position()=%s]' % str(int(index)+1))
        # save reportname for later usage
        report_name=str(element.get_attribute('innerText'))
        logger.debug(" ==> Executing report [%s] " % report_name)
        # output something in the preview on cometa front
        send_step_details(context, '%d/%d - Testing report %s' % (int(index+1), total, str(element.get_attribute('innerText'))))
        # Open report
        logger.debug("Clicking on report")
        context.browser.execute_script("arguments[0].click();", element)
        # Wait for iframe or report content
        logger.debug("Waiting for IBM Cognos iFrame to appear")
        waitSelector(context, 'css', 'iframe:nth-child(2), .ViewerContainer iframe, .pageViewContent')
        # Try to switch to new iframe if exists (only reports with prompts have iframe)
        had_iframe = False
        iframe = context.browser.find_elements(By.CSS_SELECTOR,  'iframe:nth-child(2), .ViewerContainer iframe' )
        if len(iframe) > 0:
            logger.debug('Switched to iframe')
            context.browser.switch_to.frame( iframe[0] )
            had_iframe = True
        # Automatically fill all necessary prompts
        logger.debug("Check for filling prompt with magic")
        if auto_select_cognos_prompts(context, parameters=parameters):
            # promptPage was filled with success
            logger.debug("Prompt Magic return true - which somehow means success")
        else:
            # promptPage failed somehow
            logger.debug("Prompt Magic returned false - which means, we should fail this report [%s]." % report_name)
            logger.info("You might want to look at the screenshot or video and adjust timeouts or fix a broken report.")
            logger.debug("Saveing report timing as step result to database.")
            over_all_results_ok=False

            # save to database
            save_message="I can test current IBM Cognos folder > Report: %s failed" % report_name
            context.CURRENT_STEP_STATUS = "Failed"
            saveToDatabase(save_message, (time.time() - report_start_time) * 1000, 0, False, context )

            # switch content
            logger.debug("Try switching to default content")
            context.browser.switch_to.default_content()

            # try closing the report view
            logger.debug("Trying to close report view")
            close_ibm_cognos_view(context, parameters)

            # finally continue
            logger.debug("Continue to next report as further testing on this report is not very useful.")
            logger.debug(" <== Report [%s] DONE " % report_name)
            continue

        # time.sleep(3)
        # Check if inner text has something
        if had_iframe:
            logger.debug("Waiting for iFrame to render IBM Cognos content")
            wait_for_element_text(context, 'body.viewer table, .clsViewerPage')
            logger.debug("Try switching to default content")
            context.browser.switch_to.default_content()
        else:
            logger.debug("Waiting for IBM Cognos pageViewContent")
            wait_for_element_text(context, '.pageViewContent')
        #
        # Save this report test to database, so the user can see in testresults the execution, screenshot and timings
        #
        logger.debug("Saveing report timing as step result to database")
        save_message="I can test current IBM Cognos folder > Report: %s tested" % report_name
        context.CURRENT_STEP_STATUS = "Success"
        saveToDatabase(save_message, (time.time() - report_start_time) * 1000, 0, True, context )

        # close the ibm cognos view
        close_ibm_cognos_view(context, parameters)
        logger.debug(" <== Report [%s] DONE " % report_name)

    # Finally, check if over_all_results_ok is false
    # .... this means, we had some problems inside the report execution loop
    # .... so we fail the overall test
    if over_all_results_ok:
        logger.debug("Overall results ok - all reports finished well.")
    else:
        logger.debug("Overall results failed - some reports have errors.")
        raise CustomError('Execution of reports in folder finish with some errors')

#
# FIXME documentation
#
def test_folder_aso(context, parameters = {}):

    # Make sure Team Folder is opened
    open_team_folder(context)
    send_step_details(context, 'Opening team folder')

    # Make sure all folder items are in the list
    send_step_details(context, 'Listing all folder items')
    cognos_scroll_folder_till_bottom(context, True)

    # Get all reports in current folder
    elements = waitSelector(context, "xpath", '//div[@id="teamFoldersSlideoutContent"]/descendant::div[@title and @class="nameColumnDiv contentListFocusable clickable active"]')

    # Get only 5 first elements for testing
    # elements = elements[:5]
    total = len(elements)
    if len(elements) == 0:
        raise CustomError('Empty folder')

    # Test every report by clicking on it, run it, check content, and finally close
    logger.debug("Looping over folder content with %d elements" % len(elements))

    # Set the over_results_ok_variable to true
    # ... if one report fails, then this will change to false
    over_all_results_ok=True

    # Loop over elements in folder
    for index, el in enumerate(elements):
        fail=False # if we want to fail faster
        print(' ==> Testing report #%s' % str(index))
        logger.debug(' ==================')
        logger.debug(' ==> Testing report #%s of %s' % (str(int(index)+1),str(total)))
        logger.debug(' ==================')
        report_start_time=time.time()
        # Make sure Team Folder is opened
        logger.debug("makeing sure team folder is opened")
        open_team_folder(context)
        # Make sure all folder items are in the list
        logger.debug("Scrolling folder to bottom")
        cognos_scroll_folder_till_bottom(context, True)
        # Retrieve element again as DOM can later be changed
        logger.debug("Getting DOM element %d in folder" % (int(index)+1))
        element = context.browser.find_element(By.XPATH, '//div[@id="teamFoldersSlideoutContent"]/descendant::div[@title and @class="nameColumnDiv contentListFocusable clickable active"][position()=%s]' % str(int(index)+1))
        # save reportname for later usage
        report_name=str(element.get_attribute('innerText'))
        logger.debug(" ==> Executing report [%s] " % report_name)
        # output something in the preview on cometa front
        send_step_details(context, '%d/%d - Testing report %s' % (int(index+1), total, str(element.get_attribute('innerText'))))
        # Open report
        logger.debug("Clicking on report")
        context.browser.execute_script("arguments[0].click();", element)

        # get the element that contains the storeId
        storeIDElement = waitSelector(context, "xpath", "//div[contains(@id, 'i') and @class='pageView']")[0]
        storeID = storeIDElement.get_attribute("id")
        logger.info("Found storeID for the report: %s" % storeID)

        # wait for report iframe element using the storeID as a name attribute
        logger.debug("Waiting for IBM Cognos iFrame to appear")
        # Try to switch to new iframe if exists (only reports with prompts have iframe)
        had_iframe = False
        try:
            iframe = waitSelector(context, "xpath", ("//iframe[@name='%s']" % storeID))
        except:
            iframe = []
        if len(iframe) > 0:
            logger.debug('Switched to iframe')
            context.browser.switch_to.frame( iframe[0] )
            had_iframe = True
        else:
            # check if there is a loader displayed
            loaderElements = context.browser.find_elements(By.CSS_SELECTOR, "table.rsBlockerDlg")
            logger.debug(loaderElements)
            if len(loaderElements) > 0 and loaderElements[0].is_displayed():
                logger.debug("%s with storeID %s took longer than timeout to load." % (report_name, storeID))
                fail = True
            else:
                logger.debug("No iFrame found and no loader found, maybe report already failed or report is fully displayed...")

        # wait for viewer
        time.sleep(5)
        # Automatically fill all necessary prompts
        logger.debug("Check for filling prompt with magic")
        if not fail and auto_select_cognos_prompts_aso(context, parameters=parameters):
            # promptPage was filled with success
            logger.debug("Prompt Magic return true - which somehow means success")
        else:
            if not fail:
                # promptPage failed somehow
                logger.debug("Prompt Magic returned false - which means, we should fail this report [%s]." % report_name)
                logger.info("You might want to look at the screenshot or video and adjust timeouts or fix a broken report.")
                logger.debug("Saving report timing as step result to database.")
            
            over_all_results_ok=False

            # save to database
            save_message="I can test current IBM Cognos folder > Report: %s failed" % report_name
            context.CURRENT_STEP_STATUS = "Failed"
            saveToDatabase(save_message, (time.time() - report_start_time) * 1000, 0, False, context )

            # switch content
            logger.debug("Try switching to default content")
            context.browser.switch_to.default_content()

            # try closing the report view
            logger.debug("Trying to close report view")
            close_ibm_cognos_view(context, parameters)

            # finally continue
            logger.debug("Continue to next report as further testing on this report is not very useful.")
            logger.debug(" <== Report [%s] DONE " % report_name)
            continue

        # time.sleep(3)
        # Check if inner text has something
        if had_iframe:
            logger.debug("Waiting for iFrame to render IBM Cognos content")
            wait_for_element_text(context, 'body.viewer table, .clsViewerPage')
            logger.debug("Try switching to default content")
            context.browser.switch_to.default_content()
        # else:
            # logger.debug("Waiting for IBM Cognos pageViewContent")
            # wait_for_element_text(context, '.pageViewContent')
        #
        # Save this report test to database, so the user can see in testresults the execution, screenshot and timings
        #
        logger.debug("Saveing report timing as step result to database")
        save_message="I can test current IBM Cognos folder > Report: %s tested" % report_name
        context.CURRENT_STEP_STATUS = "Success"
        saveToDatabase(save_message, (time.time() - report_start_time) * 1000, 0, True, context )

        # close the ibm cognos view
        close_ibm_cognos_view(context, parameters)
        logger.debug(" <== Report [%s] DONE " % report_name)

    # Finally, check if over_all_results_ok is false
    # .... this means, we had some problems inside the report execution loop
    # .... so we fail the overall test
    if over_all_results_ok:
        logger.debug("Overall results ok - all reports finished well.")
    else:
        logger.debug("Overall results failed - some reprots have errors.")
        raise CustomError('Execution of reports in folder finish with some errors')

# Closes the first IBM Cognos View in the drop down selector
# Example: I can close the current IBM Cognos report view "Sales Report Q4"
@step(u'I can close the current IBM Cognos report view "{parameters}"')
@done(u'I can close the current IBM Cognos report view "{parameters}"')
def step_impl(context, parameters = {}):
    close_ibm_cognos_view(context, parameters)

def close_ibm_cognos_view(context, parameters = {}):
    # Close current report view
    logger.debug("Closing current report content")
    elements = waitSelector(context, 'css', '[id="com\.ibm\.bi\.glass\.common\.viewSwitcher"]')
    context.browser.execute_script("arguments[0].click();", elements[0])
    elements = waitSelector(context, 'css', '.popover.switcher ul li:last-child .removeItemIcon')
    elements[0].click()
    time.sleep(1)
    elements = context.browser.find_elements(By.CSS_SELECTOR, '.button.dialogButton[id=ok]')
    if len(elements) > 0:
        elements[0].click()
    time.sleep(0.5)


# Allows to test all reports inside a folder of a Cognos installation with {key:value} parameters to autfill promptPages. Example: "PE|PeriodID:1269;CO:01"
# Example: I can test current IBM Cognos folder using parameters "PE|PeriodID:1269;CO:01"
@step(u'I can test current IBM Cognos folder using parameters "{parameters}"')
@done(u'I can test current IBM Cognos folder using parameters "{parameters}"')
def step_impl(context, parameters):
    # Load parameters
    parameters = load_parameters(parameters)
    test_folder_aso(context, parameters)

# FIX ME .... make this MIF unspecific
# Test if can access to a folder relative to the root directory of the URL specified
@step(u'I can test the folder "{foldername}"')
@done(u'I can test the folder "{foldername}"')
def step_impl(context, foldername):
    els = waitSelector(context, "css", '.cctable a[href*="3_5f8"]')
    logger.debug("Total number of folder elements to test: ", len(els) )
    no_of_link=1
    for link in els:
        # prepare the URL
        try:
            logger.debug(no_of_link,".) Link:",link.get_attribute("href"))
        except:
            sleep(1)
            logger.debug(no_of_link,".) Link:",link.get_attribute("href"))
        no_of_link=no_of_link+1
        actlink= link.get_attribute("href").replace("run.prompt=true","run.prompt=false")
        actlink= actlink + "&p_COND=xxxFIXMExxx&p_DATE_F=2017-10-01&p_AA=01&p_AA_SCO=12&p_AC=0500000000&p_ADT=ALL&p_APV=01&&p_AS=01&p_CO=1230&p_CU=GBP&p_DC=_&p_ET=ALL&p_EVT=01&p_MU=1110&p_PE=0361&p_PE1=0361&p_PE2=0304&p_PE3=0355&p_PE4=0329&p_PE5=0363&p_PNR_F=0010000&p_PNR_T=7010000&p_PT=104030&p_SNR_F=000&p_SNR_T=999&p_TOL=000000001"

        # open new tab with URL
        logger.debug("Will open window-URL: ", actlink)
        current_value = old_value = context.browser.window_handles
        context.browser.execute_script("window.open('"+actlink+"');")
        looper=0

        # wait for page to be ready
        while len(current_value) == len(old_value) and looper < 200:
            time.sleep(0.1)
            current_value = context.browser.window_handles
            looper=looper+1
            context.browser.switch_to.window(context.browser.window_handles[-1])

        try:
            element = WebDriverWait(context.browser, 30).until(
                EC.presence_of_element_located((By.ID, 'CVReport_NS_'))
            )
        except:
            element = WebDriverWait(context.browser, 10).until(
                EC.presence_of_element_located((By.ID, 'CVReport_NS_'))
            )
        # additional sleep ... better further intelligent wait TODO
        time.sleep(1.5)
        # take a screenshot
        context.SCREENSHOT_PREFIX = context.active_environment[0]+"_"+context.active_environment[1]+"_"+foldername+"_AL_"+str(no_of_link)+"_"
        
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S_%f")  # Format: YYYYMMDD_HHMMSS_microseconds
        screenshots_step_path = os.path.join(context.SCREENSHOTS_PATH, timestamp)
        logger.debug(f"step_screenshot_path : {screenshots_step_path} ")
        screenshot_data = takeScreenshot(context.browser, screenshots_step_path)
        toWebP_from_data(screenshot_data, screenshots_step_path)

        # see if style is ready
        context.STYLE_IMAGE = context.SCREENSHOT_PATH + 'styles/' + SCREENSHOT_PREFIX + context.SCREENSHOT_PREFIX + '.png'
        if os.path.isfile(context.STYLE_IMAGE):
            logger.debug("Style image is present:"+context.STYLE_IMAGE)
        else:
            # not ready ... copy actual screenshot as style
            logger.debug("Saveing as style")
            os.system("cp -rp "+context.COMPARE_IMAGE+" "+context.STYLE_IMAGE)

        # compare
        context.DIFF_IMAGE = context.SCREENSHOT_PATH+'diffs/'+context.SCREENSHOT_FILE+'_diff.png'
        compareImage(context)

        # close tab and switch back to previous window-handle
        logger.debug("----> done ... closing window now")
        time.sleep(1.5)
        context.browser.close()
        time.sleep(0.5)
        context.browser.switch_to.window(context.browser.window_handles[0])
        time.sleep(0.5)

# Change Language depending on the version of Cognos
# @step(u'Change Cognos Language in versión "{cognosVersion}"')
# @done(u'Change Cognos Language in versión "{cognosVersion}"')
# def step_impl(context,browserTitle):
    # Detect Environment

    # Execute the steps in feature 33 by changing the environment-dependent settings.


# Checks if the current Tab Title is/contains some sentence
# Example: BrowserTitle is "Home - Complete Meta Test Platform"
@step(u'BrowserTitle is "{browserTitle}"')
@done(u'BrowserTitle is "{browserTitle}"')
def step_impl(context,browserTitle):
    logger.debug("Checking browser.title for [%s]" % browserTitle)
    logger.debug("Browser title is: [%s]" % context.browser.title.strip())
    if context.browser.title.strip() != browserTitle:
        raise CustomError("Did not find the browser title specified but found %s." % (context.browser.title.strip()))

# Closes the browser and reverts to latest opened tab/window if available
# Example: Close the browser
@step(u'Close the browser')
@done(u'Close the browser')
def step_impl(context):
    context.browser.quit()

# Reloads the current page
# Example: Reload page
@step(u'Reload page')
@done(u'Reload page')
def step_impl(context):
    context.browser.refresh()

# Scrolls the page to a given amount of pixels in the Y axis
# Example: Scroll to "30"px
@step(u'Scroll to "{amount}"px')
@done(u'Scroll to "{amount}"px')
def step_iml(context, amount):
	context.browser.execute_script("window.scrollTo(0,"+str(amount)+")")

# Scrolls to a given amount of pixels in the Y axis inside a specific element using a CSS selector
# Example: Scroll to "200"px on element "//div[@class='welcome_logo']"
@step(u'Scroll to "{amount}"px on element "{selector}"')
@done(u'Scroll to "{amount}"px on element "{selector}"')
def step_iml(context, amount, selector):
    elements = waitSelector(context, "css", selector)
    context.browser.execute_script("arguments[0].scrollTo(0,%s)" % amount, elements[0])

# Scrolls to a given amount of pixels in the Y axis inside a specific element using a CSS selector
# Example: Scroll element "//div[@class='dual-scrollable-container']" by "200"px on x-axis and "-50"px on y-axis
@step(u'Scroll element "{selector}" by "{x_amount}"px on x-axis and "{y_amount}"px on y-axis')
@done(u'Scroll element "{selector}" by "{x_amount}"px on x-axis and "{y_amount}"px on y-axis')
def step_iml(context, x_amount, y_amount, selector):
    logger.debug(f"Scrolling element {selector} by {x_amount}px on x-axis and {y_amount}px on y-axis")
    logger.debug(f"Searching for element {selector} ")
    elements = waitSelector(context, "css", selector)
    logger.debug(f"Element found: {elements}")
    context.browser.execute_script("arguments[0].scrollBy(%s,%s)" % (x_amount, y_amount), elements[0])
    logger.debug(f"Element has been scrolled")

# Set a value on an element, normally used for inputs
# Example: Set value "Rock&Roll" on "(//input[@formcontrolname="address_to_add"])[1]"
@step(u'Set value "{text}" on "{selector}"')
@done(u'Set value "{text}" on "{selector}"')
def step_iml(context, text, selector):
    send_step_details(context, 'Looking for selector')
    element = waitSelector(context, "css", selector)
    for i in range(0, 10):
        try:
            elementInteractable = WebDriverWait(context.browser, 10).until(CEC.element_to_be_interactable(element[0]))
            if elementInteractable:
                send_step_details(context, 'Setting value')
                element[0].send_keys(text)
                send_step_details(context, 'Checking if the value is set')
                valueSet = WebDriverWait(context.browser, 10).until(CEC.text_to_be_present_in_element_value(element[0], text))
                if valueSet:
                    return True
        except ElementNotInteractableException as err:
            logger.error("Element is not interactable yet will wait.")
            time.sleep(1)
        except TimeoutException as err:
            logger.error("Element was not clickable or value was unable to set, will try again.")
    raise CustomError("Unable to set set the value, maybe there is another element in front?")

# Set a value on an element, normally used for inputs but when user do not want to verify the value from input box
# Example: Set value "Rock&Roll" on "(//input[@formcontrolname="address_to_add"])[1]" and do not check the value
@step(u'Set value "{text}" on "{selector}" and do not check the value')
@done(u'Set value "{text}" on "{selector}" and do not check the value')
def step_iml(context, text, selector):
    send_step_details(context, 'Looking for selector')
    element = waitSelector(context, "css", selector)
    try:
        elementInteractable = WebDriverWait(context.browser, 10).until(CEC.element_to_be_interactable(element[0]))
        if elementInteractable:
            send_step_details(context, 'Setting value')
            element[0].send_keys(text)
            return True
    except ElementNotInteractableException as err:
        logger.error("Element is not interactable yet will wait.")
    except TimeoutException as err:
        logger.error("Element was not interactable or value was unable to set, will try again.")

# Send any keys, this simulates the keys pressed by the keyboard
# Example: Send keys "F5"
@step(u'Send keys "{keys}"')
@done(u'Send keys "{keys}"')
def step_iml(context, keys):
    elem=context.browser.switch_to.active_element
    description = context.text
    if description is not None and description:
        # User provided a multiline value, use it instead
        keys = description
    upperKey = str(keys).upper()
    if hasattr(Keys, upperKey):
        elem.send_keys(getattr(Keys, upperKey))
    else:
        elem.send_keys(keys)


@step(u'Clear value on "{selector}"')
@done(u'Clear value on "{selector}"')
def step_clear_textbox(context, selector):
    send_step_details(context, 'Looking for selector')    
    element = waitSelector(context, "css", selector)
    for i in range(0, 10):
        try:
            elementInteractable = WebDriverWait(context.browser, 10).until(CEC.element_to_be_interactable(element[0]))
            if elementInteractable:
                send_step_details(context, 'Clearing the text box')
                element[0].clear()
                # Optionally, verify the box is empty
                valueCleared = WebDriverWait(context.browser, 10).until(
                    lambda driver: element[0].get_attribute("value") == ""
                )
                if valueCleared:
                    return True
        except ElementNotInteractableException:
            logger.error("Element is not interactable yet, will wait.")
            time.sleep(1)
        except TimeoutException:
            logger.error("Element was not clickable or value was unable to clear, will try again.")
    raise CustomError("Unable to clear the value, maybe there is another element in front?")


# Focus on element using a CSS selector
# Example: Focus on element with "//div[contains(@class, 'search-container')]"
@step(u'Focus on element with "{css_selector}"')
@done(u'Focus on element with "{css_selector}"')
def step_iml(context, css_selector):
    send_step_details(context, 'Looking for selector')
    waitSelector(context, "css", css_selector)
    send_step_details(context, 'Focusing on element')
    context.browser.execute_script('let elem=document.querySelector("'+css_selector+'"); elem.scrollIntoView(); elem.focus()')

# Highlight element
# Example: Highlight element with "//div[contains(@class, 'search-container')]"
@step(u'Highlight element with "{selector}"')
@done(u'Highlight element with "{selector}"')
def step_iml(context, selector):
    send_step_details(context, 'Looking for selector')
    element = waitSelector(context, "css", selector)
    if isinstance(element, list) and len(element) > 0:
        element = element[0]
    send_step_details(context, 'Highlighting the element')
    context.browser.execute_script('''
    const element = arguments[0];
    if (!window.cometa) window.cometa = {}
    window.cometa.oldOutlineValue = element.style.outline;
    window.cometa.oldOutlineOffsetValue = element.style.outlineOffset;
    element.style.outline = "2px solid #f00";
    element.style.outlineOffset = "1px";
    ''', element)

    # set highlight setting to be removed after step
    context.highlighted_element = {
        'element': element
    }

# Highlight element
# Example: Highlight "//span[text()='Home']" on the page
@step(u'Highlight "{text}" on the page')
@done(u'Highlight "{text}" on the page')
def step_iml(context, text):
    send_step_details(context, 'Highlighting the text')
    # action = ActionChains(context.browser)
    # action.key_down(Keys.LEFT_CONTROL).send_keys('f').key_up(Keys.LEFT_CONTROL).perform()
    context.browser.execute_script(f'''
    window.find('{text}', false, false, true);
    ''')
    # context.browser.execute_script(f'''
    # const element = document.body;
    # const pattern = new RegExp('({text})', 'gi');
    # element.innerHTML = element.innerHTML.replaceAll(
    #     pattern,
    #     "<mark data-by='co.meta'>$1</mark>"
    # );
    # ''')

    # # set highlight setting to be removed after step
    context.highlighted_text = True

# Press Enter key
# Example: Press Enter
@step(u'Press Enter')
@done(u'Press Enter')
def step_iml(context):
    elem=context.browser.switch_to.active_element
    elem.send_keys(Keys.RETURN)

# Press Tab key
# Example: Press TAB
@step(u'Press TAB')
@done(u'Press TAB')
def step_iml(context):
    elem=context.browser.switch_to.active_element
    elem.send_keys(Keys.TAB)

# Presses a set of keys sent by the user as a parameter. If the keys are separated by '+', press them simultaneosly. If they are
# separated by ';', start pressing the keys once the previous set is released
# If the key combination modifies the browser state (create new tab, close browser...), the key combination won't work
# Example: Press the following set of keys "CTRL+S;ALT+F4"
@step(u'Press the following set of keys "{keySet}"')
@done(u'Press the following set of keys "{keySet}"')
def step_impl(context, keySet):
    logger.debug('Executing action of pressing a set of keys sent by the user')
    # Removes all the spaces from the string and splits it into independent press key sets
    independentKeyArray = keySet.strip().split(';')
    # Create an array for simultaneous press key set
    simultaneousKeyArray = []
    # Loops over each independent press key set
    for i in independentKeyArray:
        keys = 'actions'
        keysUp = ''
        # Capitalizes and splits each index into simultaneous press key sets
        for key in i.split('+'):
            key = key.strip()
            if key.upper() in dir(Keys):
                # Code transformation for special characters
                keys = keys + '.key_down(Keys.%s)' % key.upper()
                keysUp = keysUp + '.key_up(Keys.%s)' % key.upper()
            else:
                # Code transformation for normal characters
                keys = keys + ".send_keys('%s')" % key.replace('\'', '\\\'')
        # Concats the key down and key up action
        keys = keys + keysUp
        logger.debug('Sending the following keys ' + keys)
        actions = ActionChains(context.browser)
        # Sends the trigger to press the keys
        eval(keys).perform()

# Checks if the current source code contains something, is case sensitive!
# Example: I can see "Home" on page
@step(u'I can see "{something}" on page')
@done(u'I can see "{something}" on page')
def step_impl(context,something):
    if not waitFor(context, something):
        raise CustomError("Unable to find %s on page." % something)


# Checks if the current source code contains a link with the desired text, is case sensitive!
# Example: I can see a link with "Home"
@step(u'I can see a link with "{linktext}"')
@done(u'I can see a link with "{linktext}"')
def step_impl(context,linktext):
    if not waitSelector(context, "link_text", linktext):
        raise CustomError("Unable to find link with %s." % linktext)

# Switches to another existing (or just created) Window/Tab
# Example: I can switch to new Window
@step(u'I can switch to new Window')
@done(u'I can switch to new Window')
def step_impl(context):
    context.browser.switch_to.window(context.browser.window_handles[-1])

# Switches to the main Window/Tab
# Example: I can switch to main Window
@step(u'I can switch to main Window')
@done(u'I can switch to main Window')
def step_impl(context):
    context.browser.switch_to.window(context.browser.window_handles[0])

# Switches to a iframe tag inside the document within the specified ID
# Example: I can switch to iFrame with id '1'
@step(u'I can switch to iFrame with id "{iframe_id}"')
@done(u'I can switch to iFrame with id "{iframe_id}"')
def step_impl(context, iframe_id):
    start_time = time.time()
    while time.time() - start_time < context.step_data['timeout']:
        logger.debug("Switching to iFrame with id: %s" % iframe_id)
        try:
            # Try getting iframe by iframe_id as text
            iframe = context.browser.find_element(By.ID,  iframe_id )
            context.browser.switch_to.frame( iframe )
            return True
        except Exception as e:
            logger.debug(e)
            import traceback
            traceback.print_exc()
            # failed switching to ID .. try overloading
            # Added method overloading with error handling by Ralf
            # Get total iframe count on current page source
            iframes = context.browser.find_elements(By.TAG_NAME, 'iframe')
            iframes_count = len(iframes)
            if isINT(iframe_id):
                iframe_id = int(iframe_id) - 1 # since the index for iFrames ids starts at 0
                if iframe_id <= iframes_count:
                    # Try getting iframe by iframe_id as index (int)
                    context.browser.switch_to.frame( iframes[iframe_id] )
                    return True

        time.sleep(0.5)
    
    raise CustomError("Unable to switch to iFrame with id: %s" % iframe_id)

# Switches to a iframe tag inside the document within the specified ID
# Example: I can switch to iFrame with selector '//iframe[@name="test_frame"]'
@step(u'I can switch to iFrame with selector "{selector}"')
@done(u'I can switch to iFrame with selector "{selector}"')
# @timeout("Waited for <seconds> seconds but was unable to find specified iFrame element.")
def step_impl(context, selector):
    send_step_details(context, f'Looking for {selector}')
    # Try getting iframe by iframe_id as text
    iframe = waitSelector(context, "name", selector)
    context.browser.switch_to.frame(iframe)


# Switches to an iframe tag inside the document within the specified ID
# Example: I can switch to iFrame with id '__privateStripeMetricsController8240'
@step(u'I can switch to iFrame with name "{iframe_name}"')
@done(u'I can switch to iFrame with name "{iframe_name}"')
def step_impl(context,iframe_name):
    send_step_details(context, 'Looking for selector')
    iframe = waitSelector(context, "name", iframe_name )
    send_step_details(context, 'Switching to iframe')
    context.browser.switch_to.frame( iframe.get_attribute('name'))

# Changes the testing context to the main document in the current Tab/Window, similar to using window.top
# Example: I switch to defaultContent
@step(u'I switch to defaultContent')
@done(u'I switch to defaultContent')
def step_impl(context):
    context.browser.switch_to.default_content()

# Checks if can click on a button with the specified text or attribute text, e.g. //button[.="'+button_name+'"] | //button[@*="'+button_name+'"]
# Example: I can click on button "//button[@aria-label='Add button']"
@step(u'I can click on button "{button_name}"')
@done(u'I can click on button "{button_name}"')
def step_impl(context, button_name):
    send_step_details(context, 'Looking for button')
    elem = waitSelector(context, "xpath", '//button[.="'+button_name+'"] | //button[@*="'+button_name+'"]')
    elem[0].click()

# Checks if can click in a button with the specified title attribute text
# Example: I can click on button "//button[@title='Add_button']"
@step(u'I can click on button with title "{button_title}"')
@done(u'I can click on button with title "{button_title}"')
def step_impl(context, button_title):
    send_step_details(context, 'Looking for button')
    elem = waitSelector(context, "css", 'button[title^="'+button_title+'"]')
    elem[0].click()

# Checks if it can click on an element using a CSS Selector
# Example: I can click on element with css selector "//button[@title='Add_button']"
@step(u'I can click on element with css selector "{css_selector}"')
@done(u'I can click on element with css selector "{css_selector}"')
def step_impl(context, css_selector):
    send_step_details(context, 'Looking for selector')
    click_element_by_css(context, css_selector)

# Checks if it can see an element using a CSS Selector
# Example: I can see element with css selector "//button[@title='Add_button']"
@step(u'I can see element with css selector "{css_selector}"')
@done(u'I can see element with css selector "{css_selector}"')
def step_impl(context, css_selector):
    send_step_details(context, 'Looking for selector')
    waitSelector(context, "css", css_selector)

# Checks if it cannot see an element using a CSS Selector until timeout
# Example: I cannot see element with selector "(//div[@role='none'])[200]"
@step(u'I cannot see element with selector "{selector}"')
@done(u'I cannot see element with css selector "{selector}"')
def cannot_see_selector(context, selector):
    # log general information about this step
    logger.debug("Running in Feature: %s " % context.feature_id )
    logger.debug("Step data: %s " % str(context.step_data) )
    timeout = context.step_data['timeout']
    logger.debug("Timeout is %s " % timeout)

    try:
        waitSelector(context, "css", selector)
        send_step_details(context, 'Looking for selector - it was found, which is bad')
        raise CustomError("Found selector using %s. Which is bad." % selector)
    except CometaTimeoutException as err:
        logger.info("Element not found which in case is good!")

# Check if the source code in the previously selected iframe contains a link with text something
# Example: I can see a link with "More options" in iframe
@step(u'I can see a link with "{linktext}" in iframe')
@done(u'I can see a link with "{linktext}" in iframe')
# @timeout("Unable to find specified linktext inside the iFrames")
def step_impl(context,linktext):
    while True:
        for child_frame in context.browser.find_elements(By.TAG_NAME, 'iframe'):
            child_frame_name = child_frame.get_attribute('name')
            child_frame_id = child_frame.get_attribute('id')
            logger.debug("Childframe:",child_frame_name)
            context.browser.switch_to.frame(child_frame_id)
            if context.browser.find_elements(By.LINK_TEXT, linktext):
                return True
        time.sleep(1)

# Selects an option defined with index from selector index defined with number. Index and number start from 0 for first element
# Example: I use selector "2" and select option "1" for Cognos promptpage
@step(u'I use selector "{number}" and select option "{index}" for Cognos promptpage')
@done(u'I use selector "{number}" and select option "{index}" for Cognos promptpage')
def step_impl(context, index, number):
    selectCognosPrompt(context, controlIndex=number, optionIndex=index)

# Selects an option value in a select input
# Example: I select the option "en"
@step(u'I select option "{option_value}"')
@done(u'I select option "{option_value}"')
def step_impl(context, option_value):
    send_step_details(context, 'Looking for option')
    elem = waitSelector(context, "css", 'option[value="'+option_value+'"]')
    try:
        elem[0].click()
    except:
        elem[0].selected = True

# Selects an option value or index for a given select element using a CSS Selector or an index
# Example: I use selector "#dropdown-id" and select option "Default"
@step(u'I use selector "{css_selector}" and select option "{value}"')
@done(u'I use selector "{css_selector}" and select option "{value}"')
def step_impl(context, css_selector, value):
    # Preformat css_selector and value variables
    css_selector = getVariableType(css_selector)
    value = getVariableType(value)

    def element_selector(context, css_selector):
        # Get selector reference in DOM
        try:
            if isinstance(css_selector, int):
                # Change name for better understanding
                index = css_selector
                send_step_details(context, 'Looking for selector')
                elements = waitSelector(context, "tag_name", "select")
                # Index is a 1+ index, therefore subtract 1 and get the select element
                index -= 1
                element = elements[index]
            else:
                send_step_details(context, 'Looking for selector')
                elements = waitSelector(context, "css", css_selector)
                element = elements[0]
            return element
        except:
            raise CustomError("Unable to find selector using %s." % css_selector)

    element = element_selector(context, css_selector)
    send_step_details(context, 'Waiting for the element to interactable.')
    while True:
        try:
            if element.is_displayed() and element.is_enabled():
                break
        except CometaTimeoutException as err:
            raise CustomError("Selector never got enabled ... will not be able to select the value.")
        except StaleElementReferenceException:
            element = element_selector(context, css_selector)

    selector = Select(element)
    send_step_details(context, 'Selecting value/index')
    # Get value or index reference in DOM
    if isinstance(value, int):
        index = value - 1
        selector.select_by_index(index)
    else:
        if value.startswith('contains:'):
            value = value[9:]
            # Get options of current selector
            options = selector.options
            index = None
            # Iterate over each option in the selector
            for idx, option in enumerate(options):
                # Get option text as lowercase
                option_value = option.text.lower()
                logger.debug(option_value, value.lower())
                # Check if option text contains the value we are searching for
                if value.lower() in option_value:
                    index = idx
                    break
            # Verify we found some option containg value
            if index is None:
                raise CustomError("Unable to find an option which contains %s." % value)
            # Perform select by index
            selector.select_by_index(index)
        else:
            try:
                selector.select_by_value(value)
            except:
                # Retry with visible value
                selector.select_by_visible_text(value)

# Selects an option value in a select input using a CSS Selector
# Example: I can select option "Default" for "#dropdown-id"
@step(u'I can select option "{option_value}" for "{css_selector}"')
@done(u'I can select option "{option_value}" for "{css_selector}"')
def step_impl(context, option_value, css_selector):
    send_step_details(context, 'Looking for selector')
    try:
        elem = waitSelector(context, "css", css_selector + ' option[value="'+option_value+'"]')
        elem[0].selected = True
    except:
        raise CustomError("Unable to select option %s for %s." % (option_value, css_selector))

# Checks if the source code contains some text, is case sensitive!
# Example: I can see "Complete Meta Test Platform"
@step(u'I can see "{something}"')
@done(u'I can see "{something}"')
def step_impl(context,something):
    try:
        waitFor(context, something)
    except:
        raise CustomError("Unable to find %s in page source." % something)

# Do a login using Basic Auth credentials, please use variables to mask sensitive values like passwords
# Example: I can do a basic auth with username "cometaUser" and "123456"
@step(u'I can do a basic auth with username "{username}" and "{password}"')
@done(u'I can do a basic auth with username & password')
def step_impl(context, username, password):
    context.browser.send_keys(username)
    context.browser.send_keys(Keys.TAB)
    context.browser.send_keys(password)
    context.browser.send_keys(Keys.ENTER)

# Sleeps for X seconds
# Example: I sleep "60" seconds
@step(u'I sleep "{sleeptime}" seconds')
@done(u'I sleep "{sleeptime}" seconds')
def step_impl(context, sleeptime):
    context.STEP_TYPE = copy.copy(context.PREVIOUS_STEP_TYPE)
    cometa_sleep(sleeptime)

# Sleeps for X seconds
# Example: I sleep "60" seconds
@step(u'I can sleep "{sleeptime}" seconds')
@done(u'I can sleep "{sleeptime}" seconds')
def step_impl(context,sleeptime):
    context.STEP_TYPE = copy.copy(context.PREVIOUS_STEP_TYPE)
    cometa_sleep(sleeptime)

# Function to check if throwed error is of type not supported
def isCommandNotSupported(err):
    # Get error message from message key or error itself
    if isinstance(err, WebDriverException):
        message = str(err)
    elif hasattr(err, 'message'):
        message = str(err.message)
    else:
        message = str(err)
    # Declarate which commands should be ignored
    commands = ['Command not supported', 'Method has not yet been implemented']
    # Check if error message is included in ignored list
    for command in commands:
        if command in message:
            return True
    return False

# FIXME Test using a custom action with parameters on x and y
def resize_browser_to(context,x,y):
    start_time = time.time()
    context.browser.set_window_position(0, 0)
    context.browser.set_window_size(x, y)
    time.sleep(.1)
    size = context.browser.get_window_size()
    logger.debug('Expected window size: width = %spx, height = %spx' % (x, y))
    logger.debug('Resulted window size: width = %spx, height = %spx.' % (size.get('width', 'undefined'), size.get('height', 'undefined')))

# Resizes the browser window to X and Y. <br>Please select a mobile browser in feature for mobile resolutions
# Example: I resize the browser to "1920" x "1080"
@step(u'I resize the browser to "{x}" x "{y}"')
@done(u'I resize the browser to "{x}" x "{y}"')
def step_impl(context, x, y):
    try:
        resize_browser_to(context,x,y)
    except Exception as err:
        if not isCommandNotSupported(err):
            raise err

# Resizes the browser window to X and Y, also checks the window size
# Example: I resize the browser to "1920" x "1080"
@step(u'I can resize the browser to "{x}" x "{y}"')
@done(u'I can resize the browser to "{x}" x "{y}"')
def step_impl(context, x, y):
    try:
        # Resize the window
        resize_browser_to(context,x,y)
        # Loop check if window size has changed with a marginal size of 5px
        # Browserstack resizes correctly in bigger resolutions but selenoid has 1px difference
        while True:
            size = context.browser.get_window_size()
            width = size["width"]
            height = size["height"]
            if ( width < int(x) + 5 or width > int(x) - 5) and ( height < int(y) + 5 or height > int(y) - 5):
                break
    except Exception as err:
        if not isCommandNotSupported(err):
            raise err

# Maximizes the browser window
# Example: Maximize the browser
@step(u'Maximize the browser')
@done(u'Maximize the browser')
def step_impl(context):
    try:
        context.browser.maximize_window()
    except Exception as err:
        if not isCommandNotSupported(err):
            raise err

# Tries to click on an element with the specified class
# Example: I click on element with classname "//div[contains(@class, 'search-container')]"
@step(u'I click on element with classname "{classname}"')
@done(u'I click on element with classname "{classname}"')
def step_impl(context, classname):
    start_time = time.time()
    try:
        send_step_details(context, 'Looking for classname')
        elem = waitSelector(context, "class", classname)
        elem[0].click()
        return True
    except Exception as e:
        raise CustomError("Could not interact with element having classname %s ." % classname)
        logger.error(str(e))
        return False

def click_on_element(elem):
    start_time = time.time()
    try:
        elem.click()
        return True
    except Exception as e:
        logger.error(str(e))
        return False

def find_and_click_link_text(context,linktext):
    start_time = time.time()
    elem = context.browser.find_elements(By.LINK_TEXT, linktext)
    if ( len(elem)>0 ):
        if ( click_on_element(elem[0]) ):
            return True
    else:
        return False

def find_and_click_custom_css(context, linktext):
    start_time = time.time()
    elem = context.browser.find_elements(By.CSS_SELECTOR, linktext)
    if ( len(elem)>0 ):
        if ( click_on_element(elem[0]) ):
            return True
    else:
        return False

def find_and_click_link_td(context,linktext):
    start_time = time.time()
    elem = context.browser.find_elements(By.XPATH, "//td[text()='"+linktext+"']")
    if ( len(elem)>0 ):
        if ( click_on_element(elem[0]) ):
            return True
    else:
        return False

def find_and_click_link_div(context,linktext):
    start_time = time.time()
    elem = context.browser.find_elements(By.XPATH, "//div[text()='"+linktext+"']")
    if ( len(elem)>0 ):
        if ( click_on_element(elem[0]) ):
            return True
    else:
        return False

def find_and_click_link_id(context,linktext):
    start_time = time.time()
    try:
        elem = context.browser.find_element(By.ID, linktext)
        if ( click_on_element(elem) ):
            return True
        else:
            return False
    except:
        return False

# try find the element in css_selector, xpath or as a link_text
# @timeout("Unable to find specified element in <seconds> seconds.")
def find_element(context, linktext):
    counter = 0
    while True:
        if ( find_and_click_link_text(context,linktext) or find_and_click_link_td(context,linktext) or find_and_click_link_div(context,linktext) or find_and_click_link_id(context,linktext) or find_and_click_custom_css(context,linktext)):
            return True
        time.sleep(1)

# Tries to make a click on link, it can be a css selector, a text link, inside a td, or a link id
# Example: I click on "About us"
@step(u'I click on "{linktext}"')
@done(u'I click on "{linktext}"')
def step_impl(context, linktext):
    send_step_details(context, 'Looking for text link')
    find_element(context, linktext)

# Scroll till the very bottom of the current folder
def cognos_scroll_folder_till_bottom(context, resetTop = False):
    elem = waitSelector(context, 'css', '#teamFoldersSlideoutContent .dataTables_scrollBody')
    elem = elem[0]
    while True:
        # Check if folder is scrolled to bottom
        isBottom = context.browser.execute_script('return arguments[0].scrollTop >= (arguments[0].scrollHeight - arguments[0].offsetHeight)', elem)
        if isBottom:
            break
        # Scroll to bottom and wait 1 second to load new folder items
        context.browser.execute_script('arguments[0].scrollTop = arguments[0].scrollHeight', elem)
        time.sleep(1)
    if resetTop:
        # Scroll to top
        context.browser.execute_script('arguments[0].scrollTop = 0', elem)

# Scroll the opened folder to the bottom
# Example: Scroll the opened folder to the bottom
@step(u'Scroll the opened folder to the bottom')
@done(u'Scroll the opened folder to the bottom')
def step_impl(context):
    elem = waitSelector(context, "css", ".dataTables_scrollBody")
    context.browser.execute_script('document.querySelector(".dataTables_scrollBody").scrollTop = document.querySelector(".dataTables_scrollBody").scrollHeight')

# Wait until I can see something on the page, useful when sleep or loading times are unknown
# Example: wait until I can see "Complete Meta Test Platform" on page
@step(u'wait until I can see "{something}" on page')
@done(u'wait until I can see "{something}" on page')
def step_impl(context, something):
    if not waitFor(context, something):
        raise CustomError("Waited for %ds but unable to find \"%s\", be aware that search is case sensitive!" % (MAXRETRIES, something))

# This step waits until the specified element is loaded
# Example: wait until "//button[@id='submit-button']" is loaded
@step(u'wait until "{selector}" is loaded')
@done(u'wait until "{selector}" is loaded')
def step_impl(context, selector):
    waitSelector(context, "xpath", selector)


# Do a login using OIDC Authentication, please use variables to mask sensitive values like passwords
# Example: I can do a OIDC auth with username "cometaUser" and "123456" 
@step(u'I can do a OIDC auth with username "{username}" and "{password}"')
@done(u'I can do a OIDC auth with username and password')
def step_impl(context, username, password):
    send_step_details(context, 'Looking for user field')
    userid = waitSelector(context, "id", 'userid')
    userid.send_keys(username)
    userid.send_keys(Keys.ENTER)
    time.sleep(1)
    send_step_details(context, 'Looking for password field')
    passwordField = waitSelector(context, "id", 'password')
    passwordField.send_keys(password)
    passwordField.send_keys(Keys.ENTER)

# Checks if an element contains a given text inside a css-property like backgroundColor, fontSize, etc
# Example: Check if "#header" contains "rgb(255, 0, 0)" in "backgroundColor"
@step(u'Check if "{css_selector}" contains "{value}" in "{css_property}"')
@done(u'Check if "{css_selector}" contains "{value}" in "{css_property}"')
def step_impl(context, css_selector, value, css_property):
    send_step_details(context, 'Looking for selector')
    waitSelector(context, "css", css_selector)
    query="return getComputedStyle(document.querySelector('"+css_selector+"'))."+css_property+".includes('"+ value +"');"
    result=context.browser.execute_script(query)
    if result != True:
        raise CustomError("Unable to find %s with %s:%s." % (css_selector, css_property, value))

# Go to previous page
# Example: Return to the previous page
@step(u'Return to the previous page')
@done(u'Return to the previous page')
def imp(context):
    context.browser.back()

# Checks if an element contains a given text inside a js-property like innerText, innerHTML, value, etc. Use the prefix "caseInsensitve:" in the value for non-exact matching
# Example: Check if "#username-input" contains "cometaUser" in JS property "value"
@step(u'Check if "{css_selector}" contains "{value}" in JS property "{js_property}"')
@done(u'Check if "{css_selector}" contains "{value}" in JS property "{js_property}"')
def imp(context, css_selector, value, js_property):
    send_step_details(context, 'Looking for selector')
    element = waitSelector(context, "css", css_selector)
    send_step_details(context, 'Checking property')
    # get asked property from element
    propValue = element[0].get_property(js_property)
    # check if property was found in element
    if propValue is None:
        raise CustomError("Property %s not found in element." % js_property)
    caseInsensitivePrefix = "caseInsensitive:"
    if value.startswith(caseInsensitivePrefix):
        # remove prefix from value
        value = str(value)[len(caseInsensitivePrefix):].lower()
        propValue = str(propValue).lower()
    # check if element property value contains the given text (case sensitive by default)
    if not value in propValue:
        raise CustomError("Unable to find %s with %s:%s." % (css_selector, js_property, value))

# Runs another feature using it's ID in the same context, useful to import common steps in multiple features
# Example: Run feature with id "320" before continuing
@step(u'Run feature with id "{feature_id}" before continuing')
@done(u'Run feature with id "{feature_id}" before continuing')
def step_impl(context, feature_id):
    pass

# Runs another feature using it's name in the same context, useful to import common steps in multiple features
# Example: Run feature with name "cometaFeature" before continuing
@step(u'Run feature with name "{feature_name}" before continuing')
@done(u'Run feature with name "{feature_name}" before continuing')
def step_impl(context, feature_name):
    pass

# add parameter to job parameters
def addParameter(context, key, value):
    job_params = json.loads(context.PARAMETERS)
    job_params[key] = value
    context.PARAMETERS = json.dumps(job_params)

# Run a JavaScript function in the current browser context
# Example: Run Javascript function "document.body.style.backgroundColor = 'lightblue';"
@step(u'Run Javascript function "{function}"')
@done(u'Run Javascript function')
def step_impl(context, function):
    if context.browser.capabilities.get('browserName', None) != 'firefox':
        _ = context.browser.get_log('browser') # clear browser logs
    js_function = context.text
    step_timeout = context.step_data['timeout']
    context.browser.set_script_timeout(step_timeout)
    try:
        result = context.browser.execute_script("""
%s
        """ % js_function)

        addParameter(context, "js_return", result)
        context.browser.set_script_timeout(30)
    except Exception as err:
        addParameter(context, "js_return", "")
        context.browser.set_script_timeout(30)
        error = str(err).split("(Session info:")[0]
        raise CustomError(error)

# Run a JavaScript function in the current browser context
# Example: On element "//button[@id='login']" run javascript function "document.body.style.backgroundColor = 'lightblue';"
@step(u'On element "{selector}" run javascript function "{function}"')  
@done(u'On element "{selector}" run javascript function "{function}"')
def step_impl(context, selector, function):
    if context.browser.capabilities.get('browserName', None) != 'firefox':
        _ = context.browser.get_log('browser') # clear browser logs
    js_function = context.text
    step_timeout = context.step_data['timeout']
    context.browser.set_script_timeout(step_timeout)
    try:

        send_step_details(context, 'Looking for selector')
        elem = waitSelector(context, "xpath", selector)
        if not click_on_element(elem[0]):
            raise CustomError("Unable to click on element with select %s" % selector)

        result = context.browser.execute_script("""
%s
        """ % js_function , elem[0])

        addParameter(context, "js_return", result)
        context.browser.set_script_timeout(30)
    except Exception as err:
        addParameter(context, "js_return", "")
        context.browser.set_script_timeout(30)
        error = str(err).split("(Session info:")[0]
        raise CustomError(error)

# Click on element using an XPath Selector
# Example: click on element with xpath "//button[@id='login']"
@step(u'click on element with xpath "{xpath}"')
@done(u'click on element with xpath "{xpath}"')
def step_impl(context, xpath):
    send_step_details(context, 'Looking for xpath element')
    elem = waitSelector(context, "xpath", xpath)
    if not click_on_element(elem[0]):
        raise CustomError("Unable to click on element with XPATH %s" % xpath)

# Use this step when a message or popup disappears so quickly 
# Examples: 
#   1. click on "//button[text()="save"]" and assert element "//p[contains(text(),'success')]" to "appeared"
#   1. click on "//button[text()="save"]" and assert element "//p[contains(text(),'success')]" to "present"
@step(u'click on "{click_element}" and assert element "{wait_element}" to "{appeared_or_present}"')
@done(u'click on "{click_element}" and assert element "{wait_element}" to "{appeared_or_present}"')
def step_impl(context, click_element, wait_element, appeared_or_present):
    send_step_details(context, 'Looking for xpath element')
    elem = waitSelector(context, "css", click_element)
    if not click_on_element(elem[0]):
        raise CustomError("Unable to click on element %s" % click_element)
    
    if appeared_or_present not in ["present","appeared"]:
        raise CustomError("'appear_or_present' value can be 'present' or 'appeared' ")
    
    send_step_details(context, f"Validating if selector '{wait_element}' is {appeared_or_present}")
        # Use waitSelector to get the element
    element = waitSelector(context, "css", wait_element)
    if type(element) == list:
        element = element[0]
    if not element:
        raise CustomError("Element is not present")
    if appeared_or_present == 'appeared':
        send_step_details(context, f"Element is present, checking for visiblity")
        while True:
            if element.is_displayed():
               break
            time.sleep(1)
            
    send_step_details(context, f"Selector '{wait_element}' is {appeared_or_present}")
    
    
# Scroll to an element using a CSS Selector
# Example: Scroll to element with css selector "#submit-button"
@step(u'Scroll to element with css selector "{selector}"')
@done(u'Scroll to element with css selector "{selector}"')
def step_impl(context, selector):
    send_step_details(context, 'Looking for selector')
    element = waitSelector(context, "css", selector)
    context.browser.execute_script('arguments[0].scrollIntoView({block: "center"})', element[0])

# Closes the current window
# Example: I can close the window
@step(u'I can close the window')
@done(u'I can close the window')
def step_impl(context):
    logger.debug("closing window")
    time.sleep(1.5)
    context.browser.close()
    time.sleep(0.5)
    context.browser.switch_to.window(context.browser.window_handles[0])
    time.sleep(0.5)

# Throws an error with a custom message and stops feature execution
# Example: Throw an error with "This is a custom error message" and leave
@step(u'Throw an error with "{message}" and leave')
@done(u'Throw an error with "{message}" and leave')
def step_impl(context, message):
    raise CustomError(message)

# Checks if an element doesn't exist using a CSS Selector
# Example: There is no coincidence with css selector ".nonexistent-element"
@step(u'There is no coincidence with css selector "{selector}"')
@done(u'There is no coincidence with css selector "{selector}"')
def step_impl(context, selector):
    found = False
    try:
        send_step_details(context, 'Looking for selector')
        elem = waitSelector(context, "css", selector)
        found = True
    except:
        pass

    if found:
        raise CustomError("Found a coincidence with css selector for %s" % selector)

def sort_column(context, column_name, reverse=False):
    context.browser.switch_to.window(context.browser.window_handles[-1])
    iframe = waitSelector(context, "name", "reportIFrame" )
    context.browser.switch_to.frame( iframe.get_attribute('name') )
    elem = waitSelector(context, "xpath", "//td/span[text()='"+column_name+"']")
    elem[0].click()
    context.browser.switch_to.window(context.browser.window_handles[-1])
    elem = waitSelector(context, "css", "button[title='Sort']")
    elem[0].click()
    send_step_details(context, 'Requesting sorting of column - waiting for iFrame to appear')
    # wait for report to be ready
    time.sleep(2)
    iframe = waitSelector(context, "name", "reportIFrame" )
    context.browser.switch_to.frame( iframe.get_attribute('name') )
    # sleep 100ms to give browser time to react and switch to the iFrame
    time.sleep(.1)
    send_step_details(context, 'Found iFrame - waiting for content to appear')
    elem = waitSelector(context, "xpath", "//td/span[text()='"+column_name+"']")
    # Click again on sort icon, if reverse sort was selected
    send_step_details(context, 'Found content - checking on reverse sorting')
    if reverse:
        context.browser.switch_to.window(context.browser.window_handles[-1])
        elem = waitSelector(context, "css", "button[title='Sort']")
        elem[0].click()
        iframe = waitSelector(context, "name", "reportIFrame" )
        context.browser.switch_to.frame( iframe.get_attribute('name') )
        # sleep 100ms to give browser time to react and switch to the iFrame
        time.sleep(.1)
        elem = waitSelector(context, "xpath", "//td/span[text()='"+column_name+"']")
    context.browser.switch_to.window(context.browser.window_handles[-1])
    send_step_details(context, 'Sorting on this column done - looping over next column')

# Sort a QueryStudio table by a given column name
# Example: I can sort QueryStudio table column with "Date;Amount"
@step(u'I can sort QueryStudio table column with "{column_name}"')
@done(u'I can sort QueryStudio table column with "{column_name}"')
def step_impl(context, column_name):
    links=column_name.split(";")
    context.browser.switch_to.window(context.browser.window_handles[-1])
    for x in links:
        # check that x is not empty
        if not x:
            continue
        # send information via websocket to user about what is going on
        send_step_details(context, 'Sorting column %s ' % x)
        sort_column(context, x)

# Add a column name in a QueryStudio table
# Example: I can add an column with "Statistics" to QueryStudio table
@step(u'I can add an column with "{column_name}" to QueryStudio table')
@done(u'I can add an column with "{column_name}" to QueryStudio table')
def step_impl(context, column_name):
    # now we can split
    links=column_name.split(";")
    context.browser.switch_to.window(context.browser.window_handles[-1])
    # loop over columns
    for x in links:
        # check that x is not empty
        if not x:
            continue
        sort=False
        reverse=False
        keep=False
        if ":" in x:
            x_split=x.split(":")
            if x_split[1].lower() == "sort":
                sort=True
            if x_split[1].lower() == "-sort":
                sort=True
                reverse=True
            if len(x_split) > 2:
                if x_split[2] == "keep":
                    keep=True
            x=x_split[0]
        # WaitFor being able to click on column to insert
        counter=0
        while True:
            try:
                elem = context.browser.find_elements(By.XPATH, '//nobr[img[contains(@src, "icon_tree_T")]]/span/span/a[text()="' + x + '"]')
                elem[0].click()
                break
            except:
                try:
                    elem = context.browser.find_elements(By.XPATH, '//nobr[img[contains(@src, "icon_tree_I")]]/span/span/a[text()="' + x + '"]')
                    elem[0].click()
                    break
                except:
                    elem = context.browser.find_elements(By.XPATH, '//nobr[img[contains(@src, "icon_tree_L")]]/span/span/a[text()="' + x + '"]')
                    elem[0].click()
                    break
            if counter > MAXRETRIES:
                raise CustomError("Waited for %ds but unable to find %s" % (MAXRETRIES, x))
            else:
                counter+=1
                time.sleep(1)
        # click on insert button
        elem = waitSelector(context, "css", ".dialogButtonText")
        elem[0].click()

        # Sleep 2 seconds, so that cognos has time to react on the insert button click
        logger.info("=============================================")
        logger.info("Column %s waiting for report to be ready" % x )
        logger.info("=============================================")
        iframe = waitSelector(context, "name", "reportIFrame" )
        context.browser.switch_to.frame( iframe.get_attribute('name') )
        elem = waitSelector(context,"xpath", "//td/span[text()=\""+x+"\"]" )

        # handle special flag sort column
        if sort:
            sort_column(context, x,reverse)

        context.browser.switch_to.window(context.browser.window_handles[-1])

        # wait for iFrame to contain the timestamp only if column should stay in report which means the page is ready
        if not keep:
            logger.info("=============================================")
            logger.info("Keeping column waiting for report to be ready")
            logger.info("=============================================")
            iframe = waitSelector(context, "name", "reportIFrame" )
            context.browser.switch_to.frame( iframe.get_attribute('name') )
            elem = waitSelector(context, "xpath", "//td[@class='pf']")
            elem[0].click();
            context.browser.switch_to.window(context.browser.window_handles[-1])

# Add a filter name to a QueryStudio table
# Example: I can add a filter with "Region" to QueryStudio table
@step(u'I can add an filter with "{filter_name}" to QueryStudio table')
@done(u'I can add an filter with "{filter_name}" to QueryStudio table')
def step_impl(context, filter_name):
    context.browser.switch_to.window(context.browser.window_handles[-1])
    elem = waitSelector(context, "link_text", filter_name)
    elem[0].click()
    elem = waitSelector(context, "css", ".dialogButtonText")
    elem[0].click()
    iframe = waitSelector(context, "id", "dialogIFrame" )
    context.browser.switch_to.frame( iframe )
    if context.browser.find_elements(By.LINK_TEXT, "OK"):
        elem = context.browser.find_elements(By.LINK_TEXT, "OK")
        elem[0].click()
    time.sleep(5)

# Add a filter name with a value to a QueryStudio table
# Example: I can add a filter to column "Sales Region" with "North America" to QueryStudio table
@step(u'I can add an filter to column "{column_name}" with "{filter_value}" to QueryStudio table')
@done(u'I can add an filter to column "{column_name}" with "{filter_value}" to QueryStudio table')
def step_impl(context, column_name, filter_value):
    notActivated=False
    missingFilter=False
    if column_name.startswith("-"):
        notActivated=True
        column_name=column_name[1:]

    if ":" in column_name:
        name_split=column_name.split(":")
        column_name = name_split[0]
        missingFilter=True
        missing=name_split[1]
    context.browser.switch_to.window(context.browser.window_handles[-1])
    iframe = waitSelector(context, "name", "reportIFrame" )
    context.browser.switch_to.frame( iframe.get_attribute('name') )
    elem = waitSelector(context, "xpath", "//td/span[text()='"+column_name+"']")
    elem[0].click()
    context.browser.switch_to.window(context.browser.window_handles[-1])
    elem = waitSelector(context, "css", "button[title='Filter']")
    elem[0].click()
    time.sleep(2)
    iframe = waitSelector(context, "id", "dialogIFrame" )
    context.browser.switch_to.frame( iframe )
    if missingFilter:
        elem = waitSelector(context, "xpath", "//a[text()='Missing values']")
        elem[0].click()
        context.browser.execute_script('document.querySelector("[name=\'QsdlgNullOptions\'] option[value=\'' + missing + '\']").selected=true')

    # wait for the filter list of values to be displayed before selecting them
    # this just works for Filter (Pick values from a list) but do not work inside if...
    #elem = waitSelector(context, "xpath", "//div[@id='selectListpList']")

    # sleep for wait last waitSelector comented
    # peding to improve to be more inteligent
    time.sleep(5)

    # work on filter dialog and select the values required from step
    if "Filter (Search for values)" in context.browser.page_source:
        if filter_value.lower() != "none":
            filter_search = waitSelector(context, "css", '[name="_sws_pSearch_searchValue"]')[0]
            filter_search.send_keys(filter_value)
            filter_search.send_keys(Keys.ENTER)
            time.sleep(5)
            elem = waitSelector(context, "xpath", "//option[contains(@value, '"+filter_value+"')]")
            elem[0].click()
            time.sleep(2)
            elem = waitSelector(context, "css", 'button[name="insertButtonpSearch"]')
            elem[0].click()
            elem = waitSelector(context, "link_text", "OK")
            elem[0].click()
    elif "Filter (Pick values from a list)" in context.browser.page_source:
        waitFor(context, "checkBoxListpList.checkData();")
        if notActivated:
            context.browser.execute_script('document.querySelector("[name=\'oExcludeSelectedValues\'] option[value=\'true\']").selected=true')
        if filter_value.lower() != "none":
            filters=filter_value.split(";")
            for f in filters:
                context.browser.execute_script('document.querySelector("input[value=\'' + f + '\']").focus()')
                context.browser.execute_script('document.querySelector("input[value=\'' + f + '\']").click()')
    elif "Lowest value" in context.browser.page_source:
        waitFor(context, column_name) # FilterDialog shows the column we are filtering on
        filters=filter_value.split(";")
        # We always need two elements to filter ... if the user only entered one value, we set the second filter value to the same value
        if len(filters) == 1:
            filters.insert(len(filters),filter_value)
        # Find the input field
        filter_from = waitSelector(context, "css", 'input[name="_textEditBoxtextBoxFrom"]')[0]
        filter_from.send_keys(filters[0])
        filter_from = waitSelector(context, "css", 'input[name="_textEditBoxtextBoxTo"]')[0]
        filter_from.send_keys(filters[1])

    # FIXME Test if necessary
    time.sleep(5) # this waits some time ... as we had issues after clicking and selecting

    # depending in the filter situation, cognos requests one or more times "OK" ... loop 3 times for 3 seconds, searching for OK button
    timer=3
    while timer != 0:
        if context.browser.find_elements(By.LINK_TEXT, "OK"):
            elem = context.browser.find_elements(By.LINK_TEXT, "OK")
            elem[0].click()
            logger.debug("Number of OK elements found: %d with timer %d" % (len(elem), timer))
        time.sleep(1)
        timer -=1
    context.browser.switch_to.window(context.browser.window_handles[-1])
    iframe = waitSelector(context, "name", "reportIFrame" )
    context.browser.switch_to.frame( iframe )
    # Wait for filter value to appear
    logger.debug("Waiting for filter item to appear in Querystudio report with max. retries: %ds " % MAXRETRIES)
    elem = waitSelector(context, "class", "filterSubtitleStyle" )
    # not sure, if useful ..... time.sleep(5)

# Negate a filter name
# Example: I can set not to the filter "Name: cometaTest"
@step(u'I can set not to the filter "{filter_text}"')
@done(u'I can set not to the filter "{filter_text}"')
def step_impl(context, filter_text):
    context.browser.switch_to.window(context.browser.window_handles[-1])
    iframe = waitSelector(context, "name", "reportIFrame" )
    context.browser.switch_to.frame( iframe.get_attribute('name') )
    elem = waitSelector(context, "xpath", "//span[contains(text(), '"+filter_text+"')]")
    elem[0].click()
    time.sleep(2)
    context.browser.switch_to.window(context.browser.window_handles[-1])
    elem = waitSelector(context, "css", "button[title='Filter']")
    iframe = waitSelector(context, "id", "dialogIFrame" )
    context.browser.switch_to.frame( iframe )
    context.browser.execute_script('document.querySelector("[name=\'oExcludeSelectedValues\'] option[value=\'true\']").selected=true')
    elem = waitSelector(context, "css", "#executeButton")
    elem[0].click()
    time.sleep(2)

# Remove a filter name from a QueryStudio table
# Example: I can remove the filter "Region: Europe"
@step(u'I can remove the filter "{filter_text}"')
@done(u'I can remove the filter "{filter_text}"')
def step_impl(context, filter_text):
    context.browser.switch_to.window(context.browser.window_handles[-1])
    iframe = waitSelector(context, "name", "reportIFrame" )
    context.browser.switch_to.frame( iframe )
    actionChains = ActionChains(context.browser)
    elem = waitSelector(context, "xpath", "//span[contains(text(), '"+filter_text+"')]")
    actionChains.context_click(elem[0]).perform()
    elem = waitSelector(context, "css", "#Delete28")
    actionChains.click(elem[0]).perform()
    time.sleep(2)
    context.browser.switch_to.window(context.browser.window_handles[-1])

# Remove a column name from a QueryStudio table
# Example: I can delete QueryStudio table column with "Revenue;Profit"
@step(u'I can delete QueryStudio table column with "{column_name}"')
@done(u'I can delete QueryStudio table column with "{column_name}"')
def step_impl(context, column_name):
    links=column_name.split(";")
    context.browser.switch_to.window(context.browser.window_handles[-1])
    send_step_details(context, 'Looping over columns to delete')
    logger.debug("Looping over columns to delete: %s" % column_name)
    for x in links:
        # check that x is not empty
        if not x:
            continue
        iframe = waitSelector(context, "name", "reportIFrame" )
        context.browser.switch_to.frame( iframe.get_attribute('name') )
        elem = waitSelector(context, "xpath", "//td/span[text()='"+x+"']")
        elem[0].click()
        time.sleep(.1)
        context.browser.switch_to.window(context.browser.window_handles[-1])
        time.sleep(.1)
        elem = waitSelector(context, "css", "button[title='Delete']")
        elem[0].click()
        send_step_details(context, 'Instructed Cognos to delete column - now waiting for dialog frame')
        logger.debug("Instructed Cognos to delete column - now waiting for dialog frame")
        try:
            iframe = waitSelector(context, "id", "dialogIFrame" )
            context.browser.switch_to.frame( iframe )
            time.sleep(.1)
            elem = waitSelector(context, "link_text", "OK")
            elem[0].click()
        except:
            logger.debug("No dialog iFrame found")
            pass
        send_step_details(context, 'Finished delete column')
        logger.debug("Finished delete column")
        context.browser.switch_to.window(context.browser.window_handles[-1])
        time.sleep(5)

# Moves a column name before another column name
# Example: I can cut QueryStudio table column with "Profit" and paste it before column with "Sales Amount"
@step(u'I can cut QueryStudio table column with "{column_name}" and paste it before column with "{column_name_2}"')
@done(u'I can cut QueryStudio table column with "{column_name}" and paste it before column with "{column_name_2}"')
def step_impl(context, column_name, column_name_2):
    context.browser.switch_to.window(context.browser.window_handles[-1])
    iframe = waitSelector(context, "name", "reportIFrame" )
    context.browser.switch_to.frame( iframe.get_attribute('name') )
    elem = waitSelector(context, "xpath", "//td/span[text()='"+column_name+"']")
    elem[0].click()
    context.browser.switch_to.window(context.browser.window_handles[-1])
    elem = waitSelector(context, "xpath", "button[title='Cut']")
    elem[0].click()

    iframe = waitSelector(context, "name", "reportIFrame" )
    context.browser.switch_to.frame( iframe.get_attribute('name') )
    elem = waitSelector(context, "xpath", "//td/span[text()='"+column_name_2+"']")
    elem[0].click()
    context.browser.switch_to.window(context.browser.window_handles[-1])
    elem = waitSelector(context, "css", "button[title='Paste (before the selected report item)']")
    elem[0].click()

# Do a Ctrl + Click using a CSS Selector
# Example: I can Control Click at "input[type='checkbox']"
@step(u'I can Control Click at "{element}"')
@done(u'I can Control Click at "{element}"')
def step_impl(context, element):
    query="""
    let ev=new MouseEvent('click', {ctrlKey: true});
    [...document.querySelectorAll(" """ + element + """ ")].forEach((link) => {
        link.dispatchEvent(ev)
    })
    """
    context.browser.execute_script(query)

# Insert custom comments in testplans
# Example: ################ Here the comment ################
@step(u'################ {comment} ################')
@done(u'################ {comment} ################')
def step_impl(context,comment):
    logger.debug("Read a comment from testplan. Will do nothing.")
    logger.info("Comment: %s" % comment)

# Insert custom comments in testplans
# Example: #Here the comment
@step(u'#{comment}')
@done(u'#{comment}')
def step_impl(context,comment):
    logger.debug("Read a comment from testplan. Will do nothing.")
    logger.info("Comment: %s" % comment)

# Save css-selector's property value if available else gets selector's innerText and saves it as an environment variable, environment variable value has a maximum value of 255 characters
# Example: Save selector "//span[text()='input_message']" value to environment variable "submit_text"
@step(u'Save selector "{css_selector}" value to environment variable "{variable_name}"')
@done(u'Save selector "{css_selector}" value to environment variable "{variable_name}"')
def step_impl(context, css_selector, variable_name):
    # get the value from the property
    send_step_details(context, 'Looking for selector')
    element = waitSelector(context, "css", css_selector)
    send_step_details(context, 'Setting environment variable')
    element = element[0]
    # get the inner text value
    result = element.get_attribute("value") or element.get_attribute("innerText")
    # remove new lines from the resutl
    result = result.replace('\n', " ")

    # add variable
    addVariable(context, variable_name, result, save_to_step_report=True)

# Add a timestamp after the prefix to make it unique
# Example: Add a timestamp to the "order" and save it to "order_id"
@step(u'Add a timestamp to the "{prefix}" and save it to "{variable_name}"')
@done(u'Add a timestamp to the "{prefix}" and save it to "{variable_name}"')
def step_impl(context, prefix, variable_name):
    # create the unique text
    text = "%s-%.0f" % (prefix, time.time())
    addVariable(context, variable_name, text, save_to_step_report=True)

# Create a sequence of random numbers of the specified size and save it to a variable
# Example: Create a string of random "6" numbers and save to "pin_code"
@step(u'Create a string of random "{x}" numbers and save to "{variable_name}"')
@done(u'Create a string of random "{x}" numbers and save to "{variable_name}"')
def step_imp(context, x, variable_name):
    import random
    text = ""
    for i in range(0, int(x)):
        text += str(random.randint(0,9))

    addVariable(context, variable_name, text)

def downloadFileFromURL(url, dest_folder, filename):
    file_path = os.path.join(dest_folder, filename)

    r = requests.get(url, stream=True)
    if r.ok:
        logger.debug("Downloading file: %s" % urllib.parse.unquote(os.path.abspath(file_path)))
        with open(file_path, 'wb') as f:
            for chunk in r.iter_content(chunk_size=1024 * 8):
                if chunk:
                    f.write(chunk)
                    f.flush()
                    os.fsync(f.fileno())

        # we copy the very same file also to a generic filename "last_downloaded_file" maintaining suffix for convinience reasons in the same folder
        file_path2 = os.path.join(dest_folder, "last_downloaded_file"+Path(file_path).suffix)
        logger.debug("Copying file also to last_downloaded_file for convinience %s", file_path2)
        shutil.copy(file_path,file_path2)

        # sleep some time to gives discs time to sync
        time.sleep(0.25)
        return file_path
    else:  # HTTP status code 4XX/5XX
        logger.error("Download failed: status code {}\n{}".format(r.status_code, r.text))
        return None

# Upload a file by selecting the upload input field and sending the keys with the folder/filename
# Example: Upload a file by clicking on "input[type='file']" using file "samplefile.txt"
@step(u'Upload a file by clicking on "{file_input_selector}" using file "{filename}"')
@done(u'Upload a file by clicking on "{file_input_selector}" using file "{filename}"')
def step_imp(context, file_input_selector, filename):
    # save the old file detector
    old_file_detector = context.browser.file_detector
    # set the new file detector to LocalFileDetector
    context.browser.file_detector = LocalFileDetector()
    # select the upload element to send the filenames to
    elements = waitSelector(context, "css", file_input_selector)
    logger.debug("Before replacing filename: %s" % filename)
    # get the target file or files
    filename = uploadFileTarget(context, filename)
    # do some logging
    logger.debug("After replacing filename: %s" % filename)
    logger.debug("Sending filename to input field")
    # send the filename string to the input field
    if type(filename) == list:
        for file in filename:
            elements[0].send_keys(file)
    else:
        elements[0].send_keys(filename)
    # reset the file detector
    context.browser.file_detector = old_file_detector

# Attach a file from Downloads folder to the current feature-result
# Example: Attach the "report.csv" from Downloads folder to the current execution results
@step(u'Attach the "{filename}" from Downloads folder to the current execution results')
@done(u'Attach the "{filename}" from Downloads folder to the current execution results')
def step_imp(context, filename):
    logger.debug("Attaching to current execution filename: %s" % filename)
    # feature resultId
    logger.debug("Feature ResultID: %s " % os.environ['feature_result_id'])
    final_filename = [os.environ['feature_result_id']+'/'+filename]
    # attaching the file to the steps
    logger.debug("Step to attach the file to: %s " % context.counters['index'])
    context.downloadedFiles[context.counters['index']] = final_filename
    logger.debug("Attached: %s" % final_filename)
    send_step_details(context, 'Attached file to to feature result')

# Download a file and watch which file is downloaded and assign them to feature_result and step_result, linktext can be a text, css_selector or even xpath
# Example: Download a file by clicking on the "Download Report"
@step(u'Download a file by clicking on "{linktext}"')
@done(u'Download a file by clicking on "{linktext}"')
def step_imp(context, linktext):
    if context.cloud != "local":
        raise CustomError("This step does not work in browserstack, please choose local browser and try again.")

    # get already downloaded files
    downloadedFiles = sum(list(context.downloadedFiles.values()), [])
    logger.debug("Downloaded files during this feature: %s" % downloadedFiles)

    send_step_details(context, 'Preparing download')
    # get session id from browser
    session_id = str(context.browser.session_id)
    # selenoid download URL
    downloadURL = "http://cometa_selenoid:4444/download/%s/" % session_id
    logger.debug("DownloadURL: %s" % downloadURL)
    logger.debug("Selenoid preference download directory: %s" % context.downloadDirectoryOutsideSelenium)

    # find and click on the element
    send_step_details(context, 'Looking for download link element')
    elem = waitSelector(context, 'css', linktext)
    send_step_details(context, 'Clicking on download element')
    # click on element
    try:
        elem[0].click()
    except:
        elem.click()

    # timer in seconds to wait until the file is downloaded
    counter = 0
    maxTimer = 120
    send_step_details(context, 'Waiting for download file')
    # get all links
    links_from_request = []
    links = []
    while counter < maxTimer:
        # get all the files downloaded from selenoid
        retries = 0
        max_retries = 20
        #  If the Selenium download request does not give a response containing the current request entry, then retry.
        while retries<max_retries:
            retries+=1
            request = requests.get(downloadURL)
            logger.debug("Download URL: %s" % downloadURL)
            logger.debug("Request content: %s" % request.content)
            # Get all links from download URL (List of files which are downloaded or still downloading)
            links_from_request = re.findall(br'href="([^\"]+)', request.content)
            # Check if file list from response content contains new file name (Compare previously downloaded file and new downloaded files)
            logger.debug("Got file list from response content %s" % links_from_request)
            if len(links_from_request) > len(downloadedFiles):
                break
            # Sleep 300 miliseconds before retring for list of files
            time.sleep(0.3)
            logger.debug(f"Retry no : {retries} | Trying to fetch request again. response content dose not contain current download link")

        logger.debug("Links to be downloaded: %s" % links_from_request)
        # check if need to continue because files are still being downloaded
        CONTINUE=False

        links = []
        # remove already downloaded files from the links
        for link in links_from_request:
            # file link is url encoded convert it to string
            file_name =  urllib.parse.unquote(link.decode('utf-8'))
            # Add feature name and replace space with _ Because "downloadFiles" variable hold file name simmiler way
            file_name_for_comparision = "%s/%s" % (os.environ['feature_result_id'], file_name.replace(" ", "_"))
            # check file name in the downloadFiles
            if file_name_for_comparision not in downloadedFiles:
                logger.debug(f"New download file name : {file_name}")
                # Add new download file to links
                links.append(file_name)

            # check if files are still being downloaded
            if re.match(r'(?:.*?\..*?download\b|^\.com\..*)', file_name):
                # there are files still being downloaded
                logger.debug('==> %s file is still being downloaded' % file_name)
                CONTINUE=True
                break

        # check if atleast one file has been downloaded
        if len(links) > 0 and not CONTINUE:
            logger.debug("There was no file to be downloaded at the link location")
            break

        # if break does not get triggered then sum the counter and sleep for 1s
        counter += 1
        time.sleep(1)

    # get the file that has been created
    downloadedFilesInThisStep = []
    # loop over all links and download them to local folders
    for link in links:
        logger.debug("Downloading %s..." % link)
        # generate link with download path and download file
        fileURL = "%s%s" % (downloadURL, link)
        logger.debug("file URL : %s " % fileURL)

        # generate filename
        filename = link.split('/')[-1].replace(" ", "_")  # be careful with file names

        logger.debug("calling function for saveing the file %s " % filename)
        downloadFileFromURL(fileURL, context.downloadDirectoryOutsideSelenium, filename)

        # logger.debug("Exact file path till downloads dir : %s " % filename)
        complete_file_name = "%s/%s" % (os.environ['feature_result_id'], filename)
        logger.debug("Complete File name with Feature : %s " % complete_file_name)

        # Attaching file to be attached with steps
        downloadedFilesInThisStep.append(complete_file_name)

    # updated downloadedFiles in context
    logger.debug("Attaching downloaded files to feature run: %s " % downloadedFilesInThisStep)
    context.downloadedFiles[context.counters['index']] = downloadedFilesInThisStep


# Delete files from Downloads folder
# Example: Delete files matching "report_*.csv" from local download folder
@step(u'Delete files matching "{pattern}" from local download folder')
@done(u'Delete files matching "{pattern}" from local download folder')
def step_imp(context, pattern):
    if context.cloud != "local":
        raise CustomError("This step does not work in browserstack, please choose local browser and try again.")

    send_step_details(context, 'Searching file in local directory')
    # list all file matching with regex
    files = glob.glob(f"{context.downloadDirectoryOutsideSelenium}/{pattern}")
    logger.debug(f"files found using regex {pattern} : {files}")
    send_step_details(context, f'Deleting {len(files)} files from download folder')
    # In case of any IO/Permission error related to files, Count how many files are deleted
    count = 0
    try:
        # loop over all the file which are be deleted
        for file in files:
            logger.debug(f'Deleting file "{file}"...')
            # delete file
            os.remove(file)
            logger.debug(f'File "{file}" deleted')
            count+=1
    except Exception as exception:
        send_step_details(context, f'{count} files deleted')
        raise exception

# Schedule a job that runs a feature with specific key:value parameters separated by semi-colon (;) and crontab patterned schedules like "* * * * *" schedule can use <today> and <tomorrow> which are replaced dynamically
# Example: Schedule Job "DataProcessingFeature" using parameters "start_date:<today>; end_date:<tomorrow>; region:US" and crontab pattern "0 6 * * *"
@step(u'Schedule Job "{feature_name}" using parameters "{parameters}" and crontab pattern "{schedule}"')
@done(u'Schedule Job "{feature_name}" using parameters "{parameters}" and crontab pattern "{schedule}"')
def step_imp(context, feature_name, parameters, schedule):

    # convers parameters to a dict
    params = {}
    # split the string by ;
    parameters_list = parameters.split(";")
    # loop over each and add them to params dict
    for parameter in parameters_list:
        # check if parameter is empty
        if parameter == "":
            continue
        # split each parameter by :
        key, value = parameter.split(":")
        # trim whitespaces from key and value if any
        key = key.strip()
        value = value.strip()
        # add key value pair to params dict
        params[key] = value

    # schedule payload
    payload = {
        "feature": feature_name,
        "parameters": params,
        "schedule": schedule,
        "owner_id": context.PROXY_USER['user_id']
    }

    # make the request to save the object
    response = requests.post(f"{get_cometa_backend_url()}/api/schedule/", headers={"Host": "cometa.local"}, json=payload)

    # if response returns a 404 throw an error
    if response.status_code != 200:
        raise CustomError(response.text)

# Removes the schedule that executed this feature y executed manually the step is ignored
# Example: Delete schedule that executed this feature
@step(u'Delete schedule that executed this feature')
@done(u'Delete schedule that executed this feature')
def step_imp(context):
    # get the parameters
    parameters = json.loads(context.PARAMETERS)
    # get job id
    jobId = parameters.get('jobId', None)

    if jobId == None:
        logger.info("No schedule executed this feature ... Not raising error just in case executed manually.")
    else:
        # payload to send
        payload = {
            'id': jobId
        }
        # make request to remove schedule
        response = requests.delete(f"{get_cometa_backend_url()}/api/schedule/", headers={"Host": "cometa.local"}, json=payload)

        # check the response if not 200 raise exception
        if response.status_code != 200:
            raise CustomError("No jobId found. That's on us though.")

# from https://stackoverflow.com/a/60049042/18232031
def index_transform(excel_index):
    match = re.match(r"^([a-z]+)(\d+)$", excel_index.lower())
    if not match:
        raise CustomError("Invalid index")

    x_cell = -1
    for idx, char in enumerate(match.group(1)[::-1]):
        x_cell += (26 ** idx) * (ord(char) - 96)  # ord('a') == 97

    y_cell = int(match.group(2)) - 1

    return y_cell, x_cell

def updateExcel(excelFilePath, cell, value, savePath):
    # import openpyxl for excel modifications
    from openpyxl import load_workbook

    # load excel file
    wb = load_workbook(filename=excelFilePath)

    # get active sheet
    sheet = wb.active

    # modify the cell with value
    sheet[cell] = value

    # save excel file back
    wb.save(filename=excelFilePath)

    # give some time for syncing filesystem
    time.sleep(0.1)

def updateCsv(excelFilePath, cell, value, savePath):
    # importing the pandas library
    import pandas as pd

    # reading the csv file, options automatically detect delimiter and lineending
    logger.debug(f"Reading CSV: {excelFilePath}")
    df = pd.read_csv(excelFilePath, sep=None, engine='python')

    # get excel equivalent to CSV index
    indexes = index_transform(cell)

    # updating the column value/data
    logger.debug(f"Updateing file in cell {cell}")
    df.iloc[indexes[0], indexes[1]] = value

    # writing into the file
    logger.debug("Saveing updated file")
    df.to_csv(savePath, index=False)

# Edit excel or csv file and set a value to a given cell. The file is saved on the same path
# Example: Edit "sales_data.csv" and set "5000" to "B3"
@step(u'Edit "{file}" and set "{value}" to "{cell}"')
@done(u'Edit "{file}" and set "{value}" to "{cell}"')
def editFile(context, file, value, cell):
    # get file path
    filePath = uploadFileTarget(context, file)
    logger.debug("File opening: %s", filePath)

    # convert csv file to execl before continuing
    oldPath = filePath
    filePath, isCSV = CSVtoExcel(context, filePath)

    updateExcel(filePath, cell, value, filePath)

    if isCSV:
        ExcelToCSV(context, filePath, oldPath)

    updateSourceFile(context, oldPath, file)

    # give some time for syncing filesystem
    time.sleep(0.1)

# Opens excel file and tests that value is found in a given cell
# Example: Open "cometa_data.xlsx" and assert "cometaUser" is in cell "B2"
@step(u'Open "{excelfile}" and assert "{value}" is in cell "{cell}"')
@done(u'Open "{excelfile}" and assert "{value}" is in cell "{cell}"')
def editFile(context, excelfile, value, cell):
    # import openpyxl for excel modifications
    from openpyxl import load_workbook

    excelFilePath = uploadFileTarget(context, excelfile)
    logger.debug("Excel file opening: %s", excelFilePath)

    # convert csv file to execl before continuing
    oldPath = excelFilePath
    excelFilePath, isCSV = CSVtoExcel(context, excelFilePath)

    # load excel file
    wb = load_workbook(filename=excelFilePath)

    # get active sheet
    sheet = wb.active

    # assert the cell with value
    cell = sheet[cell]
    logger.debug(cell.data_type)
    cell_value = str(cell.internal_value).strip()
    value = str(value).strip()
    logger.debug("Cell value: %s" % cell_value)
    logger.debug("Value to compare: %s" % value)

    assert cell_value == value, f"Cell value ({cell_value}) does not match expected value ({value})."

# Opens excel file adds a variable to environment and sets the value as seen in Excel cell
# Example: open "cometa_results.xlsx" and set environment variable "total_fails" with value from cell "C5"
@step(u'Open "{excelfile}" and set environment variable "{variable_name}" with value from cell "{cell}"')
@done(u'Open "{excelfile}" and set environment variable "{variable_name}" with value from cell "{cell}"')
def editFile(context, excelfile, variable_name, cell):
    # import openpyxl for excel modifications
    from openpyxl import load_workbook

    excelFilePath = uploadFileTarget(context, excelfile)
    logger.debug("Excel file opening: %s", excelFilePath)

    # convert csv file to execl before continuing
    oldPath = excelFilePath
    excelFilePath, isCSV = CSVtoExcel(context, excelFilePath)

    # load excel file
    wb = load_workbook(filename=excelFilePath)

    # get active sheet
    sheet = wb.active

    # assert the cell with value
    logger.debug("Setting value value: %s to variable %s " % (sheet[cell].value, variable_name) )

    # add variable
    addVariable(context, variable_name, sheet[cell].value, save_to_step_report=True)

# get total cells from the cell ranges and
def getTotalCells(sheet, cells, values=[]):
    # split cells using semicolon (;)
    cells = cells.split(";")
    # save total cells to an array
    totalCells = []
    # loop over all the cells and check their values agaist the values index
    for cell_range in cells:
        # check if cell range contains a letter only
        # date pattern to look for
        pattern = r'^(?P<column>[A-Z]+)(?P<row>[0-9]+):\1$'
        # match pattern to the paramerters value
        match = re.search(pattern, str(cell_range))
        # if match was found
        if match:
            # get all the matches
            groups = match.groupdict()
            # update the cell range
            cell_range = "%s%s" % (cell_range, str(int(groups['row']) + len(values) - 1))
            logger.debug("New range: %s" % cell_range)

        cell = sheet[cell_range]
        if type(cell) == tuple:
            for x in cell:
                for y in x:
                    totalCells.append(y)
        else:
            totalCells.append(cell)
    return totalCells

def CSVtoExcel(context, filePath):
    NEWFILE=filePath
    ISCSV=False
    if filePath.endswith(".csv"):
        NEWFILE="%s.xlsx" % filePath
        send_step_details(context, 'Converting CSV file to Excel file.')
        import pandas as pd
        df = pd.read_csv(filePath,sep=None, engine='python')
        df.to_excel(NEWFILE, index=None)
        ISCSV=True
    send_step_details(context, '')
    return (NEWFILE, ISCSV)

def ExcelToCSV(context, filePath, newPath):
    if filePath.endswith(".xls") or filePath.endswith(".xlsx"):
        send_step_details(context, 'Converting Excel file to CSV file.')
        import pandas as pd
        df = pd.read_excel(filePath) # or other encodings
        df.to_csv(newPath, index=None)
    return newPath

# Assert values inside the excel file, generates a CSV file with the result
# Example: Open "cometa_data.xlsx" and test that cells "A1;B2;C3" contain "100;200;300" options "match exact"
@step(u'Open Excel from "{file}" and test that cells "{excel_range}" contain "{values}" options "{match_type}"')
@done(u'Open Excel from "{file}" and test that cells "{excel_range}" contain "{values}" options "{match_type}"')
def excel_step_implementation(context, file, excel_range, values, match_type):

    # match options
    match_options = ['match exact', 'match any', 'match X number of times', 'match partial']
    # check if match type is one of next options
    if match_type not in match_options:
        raise CustomError("Unknown match_type, match type can be one of these options: %s." % ", ".join(match_options))

    # import openpyxl for excel modifications
    from openpyxl import load_workbook

    excelFilePath = uploadFileTarget(context, file)
    logger.debug("Excel file opening: %s", excelFilePath)

    # check if file is a CSV file if so convert it to excel
    OLDPATH=excelFilePath
    excelFilePath, ISCSV = CSVtoExcel(context, excelFilePath)
    logger.debug(f"After CSV convert: {excelFilePath} & {ISCSV}")

    # load excel file
    wb = load_workbook(filename=excelFilePath)

    # get active sheet
    sheet = wb.active

    # make sure that the cells and values contain the same number of object
    values = values.split(";")

    # separate cells using semicolon
    cells = getTotalCells(sheet, excel_range, values)


    if len(cells) != len(values):
        raise CustomError("Cells and values should contain the same number of properties separated by semicolon (;). Total cells found %d and total values found: %d." % (len(cells), len(values)))

    # save the result in a dict to later convert it to CSV
    result = []

    # compare cell value with values provided
    for i in range(0, len(cells)):
        result.append({
            "cell": cells[i].coordinate,
            "cell_value": cells[i].value,
            "expected_value": values[i],
            "status": cells[i].value == values[i]
        })

    # logic based on match type
    allStatus = [row['status'] for row in result]
    overAllStatus = False
    if match_type == 'match exact':
        overAllStatus = not (False in allStatus)
    elif match_type == 'match partial':
        overAllStatus = True in allStatus
    else:
        raise CustomError("match_type: %s is not implemented yet." % match_type)

    # save date and time an later format it
    dateTime = datetime.datetime.now()

    # generate a csv file
    fileContent = [
        ["Feature ID", int(context.feature_id)],
        ["Feature Result ID", int(os.environ['feature_result_id'])],
        ["Step number (starts from 1)", context.counters['index'] + 1],
        ["Date & Time", dateTime.strftime("%Y-%m-%d %H:%M:%S")],
        ["Overall Status", "Passed" if overAllStatus else 'Failed'],
        ["Option", match_type],
        [""],
        ["Comparison Details:"],
        ["Cell", "Cell Value", "Expected Value", "Status"]
    ]

    # print all the rows
    for row in result:
        fileContent.append([
            row["cell"],
            row["cell_value"],
            row["expected_value"],
            row["status"]
        ])

    # save lists to file and link it to the step
    fileName = "Excel_Assert_Values_%s.csv" % dateTime.strftime("%Y%m%d%H%M%S%f")
    filePath = "%s/%s" % (context.downloadDirectoryOutsideSelenium, fileName)

    # open file and write to it
    with open(filePath, 'w+', encoding="utf_8_sig") as fileHandle:
        writer = csv.writer(fileHandle, dialect="excel", delimiter=",", lineterminator='\n')
        writer.writerows(fileContent)

    # updated downloadedFiles in context
    context.downloadedFiles[context.counters['index']] = ["%s/%s" % (os.environ['feature_result_id'], fileName)]

    if not overAllStatus:
        raise CustomError("Excel assert values failed, please view the attachment for more details.")

# Possible options: do not count empty cells or include empty cells
# Example: Open "cometa_data.xlsx" and compare the number of rows in the "B" column, starting from row "2", to ensure that there are "15" rows with option "do not count empty cells"
@step(u'Open "{excelfile}" and compare the number of rows in the "{column}" column, starting from row "{starting_row}", to ensure that there are "{total_rows}" rows with option "{option}"')
@done(u'Open "{excelfile}" and compare the number of rows in the "{column}" column, starting from row "{starting_row}", to ensure that there are "{total_rows}" rows with option "{option}"')
def assert_row_count(context, excelfile, column, starting_row, total_rows, option):
    # match options
    assert_options = ['do not count empty cells', 'include empty cells']
    # check if match type is one of next options
    if option not in assert_options:
        raise CustomError("Unknown option, option can be one of these options: %s." % ", ".join(assert_options))

    # import openpyxl for excel modifications
    from openpyxl import load_workbook

    excelFilePath = uploadFileTarget(context, excelfile)
    logger.debug("Excel file opening: %s", excelFilePath)

    # check if file is a CSV file if so convert it to excel
    OLDPATH=excelFilePath
    excelFilePath, ISCSV = CSVtoExcel(context, excelFilePath)
    logger.debug(f"After CSV convert: {excelFilePath} & {ISCSV}")

    # load excel file
    wb = load_workbook(filename=excelFilePath)

    # get active sheet
    sheet = wb.active
    # get max rows from the sheet with data
    max_rows = sheet.max_row

    # get all rows from start to finish
    cells = sheet[f"{column}{starting_row}:{column}{max_rows}"]
    totalCells = len(cells)

    if option == "do not count empty cells":
        totalCells = 0
        for cell in cells:
            try:
                cell = cell[0]
                if cell is not None and str(cell.value).strip():
                    totalCells += 1
            except Exception as err:
                logger.error("Error occurred while trying to access cell value.")
                logger.exception(err)

    logger.debug(f"Total cells found {totalCells} using option {option}.")

    assert totalCells == int(total_rows), f"The expected number of rows was {total_rows}, but the actual number of rows found using the '{option}' option was {totalCells}."

# Saves css_selectors innertext into a list variable. use "unique:<variable>" to make values distinct/unique. Using the variable in other steps means, that it includes "unique:", e.g. use "unique:colors" in other steps
# Example: Save list values in selector "//img[@class='img-blue']" and save them to variable "image_variable"
@step(u'Save list values in selector "{css_selector}" and save them to variable "{variable_name}"')
@done(u'Save list values in selector "{css_selector}" and save them to variable "{variable_name}"')
def imp(context, css_selector, variable_name):
    # get all the values from css_selector
    send_step_details(context, 'Looking for selector')
    elements = waitSelector(context, "css", css_selector)
    # get elements values
    send_step_details(context, 'Saving list')
    element_values = [element.get_attribute("innerText") or element.get_attribute("value") for element in elements]

    # check if variable name starts with "unique:", if so make elements distinct/unique (from ticket 2976)
    # TODO: implement options to be used in order to not polute the variable names
    if ( variable_name.startswith("unique") ):
        new_elements = []
        logger.debug("AMVARA: received variable starting with unique. Will order elements in list: %s." % str(element_values))
        for i in element_values:
            if i not in new_elements:
                new_elements.append(i)
        element_values = new_elements
        logger.debug("AMVARA: Ordered elements: %s." % str(element_values))

    # convert to string
    elements_list = ";".join(element_values)

    # add the variable
    addVariable(context, variable_name, elements_list, save_to_step_report=True)

# Make a request to Open Weather Map and get Weather information about specific City, using units specified at https://openweathermap.org/current and your API Key
# Example: Weather temperature from Open Weather Map for "London" with "metric" using "your_api_key" and save to variable "weather_temp"
@step(u'Get weather temperature from Open Weather Map for "{city}" with "{units}" using "{apikey}" and save to variable "{variable_name}"')
@done(u'Get weather temperature from Open Weather Map for "{city}" with "{units}" using "{apikey}" and save to variable "{variable_name}"')
def step_imp(context, city, units, apikey, variable_name):
    # make a request to openweathermap
    req = requests.get("https://api.openweathermap.org/data/2.5/weather?q=%s&appid=%s&units=%s" % (str(city), str(apikey), (str(units) or "Standard")))
    try:
        res = req.json()
        temp = res['main']['temp']
        addVariable(context, variable_name, temp)
    except Exception as error:
        raise CustomError(error)

# Compare two numbers with a variance
# Example: Compare "32" and "28" with a "5"
@step(u'Compare "{value_one}" and "{value_two}" with a "{variance}"')
@done(u'Compare "{value_one}" and "{value_two}" with a "{variance}"')
def step_imp(context, value_one, value_two, variance):
    # get number/float inside value_one
    number_one = re.findall("\d*\.?\d*?", value_one)[0]
    # get number/float inside value_two
    number_two = re.findall("\d*\.?\d*?", value_two)[0]
    # get number inside variance
    number_variance = re.findall("\d*\.?\d*?", variance)[0]

    # get the difference between both numbers
    diff = abs(float(number_one) - float(number_two))

    if int(diff) > int(number_variance):
        raise CustomError("Difference (%s) is greater than variance (%s) specified." % (str(diff), str(number_variance)))

# Generate One-Time Password (OTP) using a pairing-key
# Example: Create one-time password of "6" digits using pairing-key "base32keyhere" and save it to crypted variable "otp"
@step(u'Create one-time password of "{x}" digits using pairing-key "{value}" and save it to crypted variable "{variable_name}"')
@done(u'Create one-time password of "{x}" digits using pairing-key "{value}" and save it to crypted variable "{variable_name}"')
def step_imp(context, x, value, variable_name):
    x = x.strip()
    value = re.sub(r"\s+", "", value)
    try:
        x = int(x)
    except:
        x = 6
        send_step_details(context, 'x should be one of these numbers: 6, 7, 8. Defaulting to 6.')
        logger.warn("x should be one of these numbers: 6, 7, 8. Defaulting to 6.")

    if x not in (6, 7, 8):
        x = 6
        send_step_details(context, 'x should be one of these numbers: 6, 7, 8. Defaulting to 6.')
        logger.warn("x should be one of these numbers: 6, 7, 8. Defaulting to 6.")

    import pyotp
    totp = pyotp.TOTP(value, digits=x)
    oneTimePassword = totp.now()
    addVariable(context, variable_name, oneTimePassword, encrypted=True, save_to_step_report=True)

def test_ibm_cognos_cube(context, all_or_partial, variable_name, prefix, suffix):
    # get the variables from the context
    env_variables = json.loads(context.VARIABLES)
    # check if variable_name is in the env_variables
    index = [i for i,_ in enumerate(env_variables) if _['variable_name'] == variable_name]

    if len(index) == 0:
        # if the length is 0 then the variable was not found
        # so we can try using the value in the variable as if it is a list of values - see issue #3881
        logger.debug("Feature: %s - Will use variable as valuelist %s " % context.feature_id, variable_name)
        send_step_details(context, 'Variable does not exist, will use the variable name as value seperate by semicolon.')

    # hold csv data that will later on be writen to a file
    fileContent = []
    fileContent.append([""])
    fileContent.append(["Checking IBM Cognos Cube Dimention:"])
    fileContent.append(["Variable List", "Status"])

    # save date and time an later format it
    dateTime = datetime.datetime.now()

    # get variable value from the variables
    variable_value = env_variables[index[0]]['variable_value']
    # split it from semi-colon (;)
    values = variable_value.split(";")

    # save all the values found and not found in the report cube
    values_found = []
    values_not_found = []

    # run checks for each value in list
    for value in values:
        # add prefix and suffix to the value
        search = "%s%s%s" % (prefix, value, suffix)
        # trim the value to be searched
        search = search.strip()
        # click on search button
        element = waitSelector(context, "xpath", '//span[text()="Search..."]', 5)
        element[0].click()
        # wait for the popup window
        waitSelector(context, "xpath", '//*[text()="Keywords:"]', 5)
        # send keys once the popup window is open
        context.browser.switch_to.active_element.send_keys(search)
        # click on search
        element = waitSelector(context, "xpath", '//button//span[text()="Search"]', 5)
        element[0].click()
        # sleep 500ms for search result to appear
        time.sleep(0.5)
        # don't throw an error if not found at the end fail the step and let user know about missing values
        try:
            # look for the search value in the search tree
            waitSelector(context, "xpath", '//*[contains(@id, "Tree_Search")]//*[text()="%s"]' % search, 5)
            logger.debug("Found %s in the report cube..." % search)
            fileContent.append([search, "Found"])
            values_found.append("%s (%s)" % (value, search))
        except Exception as e:
            logger.error(str(e))
            # append the value in not founds
            values_not_found.append("%s (%s)" % (value, search))
            fileContent.append([search, "Not Found"])
        # go back and search for next value
        element = waitSelector(context, "css", "#idSourcesPane_btnClearSearch", 5)
        element[0].click()

    fileContent.insert(0, ["Feature ID", int(context.feature_id)])
    fileContent.insert(1, ["Feature Result ID", int(os.environ['feature_result_id'])])
    fileContent.insert(2, ["Step number (starts from 1)", context.counters['index'] + 1])
    fileContent.insert(3, ["Variable", variable_name])
    fileContent.insert(4, ["Date & Time", dateTime.strftime("%Y-%m-%d %H:%M:%S")])
    fileContent.insert(5, ["Overall Status", "Failed" if len(values_not_found) > 0 else 'Passed'])

    # save lists to file and link it to the step
    fileName = "IBM_cube_test_%s.csv" % dateTime.strftime("%Y%m%d%H%M%S%f")
    filePath = "%s/%s" % (context.downloadDirectoryOutsideSelenium, fileName)

    # open file and write to it
    with open(filePath, 'w+', encoding="utf_8_sig") as fileHandle:
        writer = csv.writer(fileHandle, dialect="excel", delimiter=",", lineterminator='\n')
        writer.writerows(fileContent)

    # updated downloadedFiles in context
    context.downloadedFiles[context.counters['index']] = ["%s/%s" % (os.environ['feature_result_id'], fileName)]

    # finally check if values_not_found has something if so fail step and let user know about the missing values
    if (len(values_not_found) > 0 and all_or_partial == 'all') or (len(values_found) == 0):
        missin_values = "; ".join(values_not_found)
        raise CustomError("Here are the missing values in Report Cube: %s" % missin_values)

# Compares a report cube's content to a list saved in variable
# Example: Test IBM Cognos Cube Dimension to contain all values from list variable "region_codes" use prefix "Region:" and suffix "0"
@step(u'Test IBM Cognos Cube Dimension to contain all values from list variable "{variable_name}" use prefix "{prefix}" and suffix "{suffix}"')
@done(u'Test IBM Cognos Cube Dimension to contain all values from list variable "{variable_name}" use prefix "{prefix}" and suffix "{suffix}"')
def imp(context, variable_name, prefix, suffix):
    test_ibm_cognos_cube(context, 'all', variable_name, prefix, suffix)

# Compares a report cube's content to a list saved in variable
# Example: Test IBM Cognos Cube Dimension to contain "partial" values from list variable "sales_codes" use prefix "" and suffix "-2023".
@step(u'Test IBM Cognos Cube Dimension to contain "{all_or_partial}" values from list variable "{variable_name}" use prefix "{prefix}" and suffix "{suffix}"')
@done(u'Test IBM Cognos Cube Dimension to contain "{all_or_partial}" values from list variable "{variable_name}" use prefix "{prefix}" and suffix "{suffix}"')
def imp(context, all_or_partial, variable_name, prefix, suffix):

    # check the value for all_or_partial
    if all_or_partial not in ['all', 'partial']:
        raise CustomError("all_or_partial value should be 'all' or 'partial'.")

    test_ibm_cognos_cube(context, all_or_partial, variable_name, prefix, suffix)

# Tests if a list of elements (identified by CSS selector) contains "all" or "partial" values from specified list variables
# Example: Test list of ".product-item" elements to contain "all" values from list variable "product_codes|additional_codes" use prefix "Code:" and suffix ""
@step(u'Test list of "{css_selector}" elements to contain "{all_or_partial}" values from list variable "{variable_names}" use prefix "{prefix}" and suffix "{suffix}"')
@done(u'Test list of "{css_selector}" elements to contain "{all_or_partial}" values from list variable "{variable_names}" use prefix "{prefix}" and suffix "{suffix}"')
def step_test(context, css_selector, all_or_partial, variable_names, prefix, suffix):

    # check if all_or_partial contains one or the other value if anything
    # else return error
    if all_or_partial != 'all' and all_or_partial != 'partial':
        raise CustomError("all_or_partial value must be all or parcial")

    # try getting the elements else set empty array
    try:
        # get all the values from css_selector
        elements = waitSelector(context, "css", css_selector)
    except CustomError as customError:
        elements = []
    # get elements values
    element_values = [(element.get_attribute("innerText") or element.get_attribute("value")).strip() for element in elements]

    # get the variables from the context
    env_variables = json.loads(context.VARIABLES)
    # check if variable_name is in the env_variables
    index = [i for i,_ in enumerate(env_variables) if _['variable_name'] in variable_names.split("|")]

    if len(index) == 0:
        raise CustomError("No variable found with names: %s" % ", ".join(variable_names.split("|")))

    # create a variables that will contain all the concated values from variables
    values = []
    # loop over indexes found while looking for variables
    for i in index:
        # get variable value from the variables
        variable_value = env_variables[i]['variable_value']
        # split it from semi-colon (;)
        values.extend(variable_value.split(";"))

    # add prefix and suffix to the values
    values = [("%s%s%s" % (prefix, value.strip(), suffix)).strip() for value in values]

    # equalize the lenght of both lists
    values_eq = values[:len(element_values)] if len(values) > len(element_values) else values
    element_values_eq = element_values[:len(values)] if len(element_values) > len(values) else element_values

    # sorted lists
    element_values_sorted = sorted(element_values)
    values_sorted = sorted(values)

    # equalize the lenght of both sorted lists
    values_sorted_eq = values_sorted[:len(element_values_sorted)] if len(values_sorted) > len(element_values_sorted) else values_sorted
    element_sorted_values_eq = element_values_sorted[:len(values_sorted)] if len(element_values_sorted) > len(values_sorted) else element_values_sorted

    # save date and time an later format it
    dateTime = datetime.datetime.now()

    # save lists to file and link it to the step
    fileName = "list_comparison_%s.csv" % dateTime.strftime("%Y%m%d%H%M%S%f")
    filePath = "%s/%s" % (context.downloadDirectoryOutsideSelenium, fileName)

    status = False
    fileContent = []

    if all_or_partial == 'all':

        fileContent.append([""])
        fileContent.append(["Comparison Details Unsorted:"])
        fileContent.append(["Element List", "Variable List", "Status"])

        # print element list and variable list
        for i in range(0, len(values) if len(values) >= len(element_values) else len(element_values)):
            # get value for values_eq
            try:
                val = values[i]
            except:
                val = ''
            # get value for element_values_eq
            try:
                element_val = element_values[i]
            except:
                element_val = ''
            # get the result
            if val == element_val:
                result = "Same"
            else:
                result = "Not Same"

            if val == '' or element_val == '':
                result = "Ignored"

            fileContent.append([
                element_val, # Element List value
                val, # Variable List value,
                result # result of comparison
            ])

        # add additional text to fileContent
        fileContent.append([""])
        fileContent.append(["Comparison Details Sorted:"])
        fileContent.append(["Sorted Element List", "Sorted Variable List", "Status"])

        # print sorted element list and sorted variable list
        for i in range(0, len(values_sorted) if len(values_sorted) >= len(element_values_sorted) else len(element_values_sorted)):
            # get value for values
            try:
                val = values_sorted[i]
            except:
                val = ''
            # get value for element_values_eq
            try:
                element_val = element_values_sorted[i]
            except:
                element_val = ''
            # get the result
            if val == element_val:
                result = "Same"
            else:
                result = "Not Same"

            if val == '' or element_val == '':
                result = "Ignored"

            fileContent.append([
                element_val, # Element List value
                val, # Variable List value,
                result # result of comparison
            ])

        status = (values_eq == element_values_eq) or (values_sorted_eq == element_sorted_values_eq)
    else:
        # add additional text to fileContent
        fileContent.append([""])
        fileContent.append(["Checking if variable list values are in elements list:"])
        fileContent.append(["Variable List", "In Elements List"])

        # print variable list and if they are in element list with yes/no
        atLeastOneValueInElements = False
        for value in values:
            result = "No"
            if value in element_values:
                result = "Yes"
                atLeastOneValueInElements = True
            # print result to the file
            fileContent.append([
                value, # Variable List value,
                result # result of comparison
            ])
        status = atLeastOneValueInElements


    fileContent.insert(0, ["Feature ID", int(context.feature_id)])
    fileContent.insert(1, ["Feature Result ID", int(os.environ['feature_result_id'])])
    fileContent.insert(2, ["Step number (starts from 1)", context.counters['index'] + 1])
    fileContent.insert(3, ["Selector", css_selector])
    fileContent.insert(4, ["Date & Time", dateTime.strftime("%Y-%m-%d %H:%M:%S")])
    fileContent.insert(5, ["Overall Status", "Passed" if status else 'Failed'])
    fileContent.insert(6, ["Option", all_or_partial])

    # open file and write to it
    with open(filePath, 'w+', encoding="utf_8_sig") as fileHandle:
        writer = csv.writer(fileHandle, dialect="excel", delimiter=",", lineterminator='\n')
        writer.writerows(fileContent)

    # updated downloadedFiles in context
    context.downloadedFiles[context.counters['index']] = ["%s/%s" % (os.environ['feature_result_id'], fileName)]

    if status:
        return True
    else:
        raise CustomError("Lists do not match, please check the attachment.")

# This step initiates a loop that runs a specific number of times, starting from a given index
# Example: Loop "3" times starting at "1" and do'
@step(u'Loop "{x}" times starting at "{index}" and do')
@done(u'Loop "{x}" times starting at "{index}" and do')
def step_loop(context, x, index):
    # check if there was an error during loop
    err = False
    # save current step index before continuing
    currentStepIndex = context.counters['index']
    # save substep index
    subStepIndex = 0
    # get all the sub steps from text
    steps = context.text
    logger.debug(f"Steps to be executed {steps}")
    # set context.insideLoop to true
    context.insideLoop = True
    # set executedStepsInLoop value
    context.executedStepsInLoop = 0
    # match regexp to find steps and step descriptions
    steps = list(filter(None, re.findall(r".*\n?(?:\t'''(?:.|\n)+?'''\n?)?", steps)))
    logger.debug(f"list of test_steps : {steps}")
    try:
        logger.debug("Steps: {}".format(steps))
        for i in range(int(index), int(x) + int(index)):
            logger.debug(f"With in the loop-Step {i}")
            # update subStepIndex to currentStepIndex
            subStepIndex = currentStepIndex
            # add a index variable to context.JOB_PARAMETERS
            params = json.loads(context.PARAMETERS)
            params['index'] = i
            context.PARAMETERS = json.dumps(params)
            for step in steps:
                if context.continue_loop:
                    logger.debug("Loop is in continue state")
                    continue
                logger.debug("Executing step: {}".format(step))
                # update steps executed
                context.executedStepsInLoop += 1
                # update subStepIndex with +1 on each step
                subStepIndex = subStepIndex + 1
                # update stepIndex with subStepIndex
                context.counters['index'] = subStepIndex
                # replace ''' to """
                step = step.replace("'''", "\"\"\"")
                # print some information
                send_step_details(context, "Executing step '%s' inside loop." % step.split('\n')[0])
                # execute the step
                context.execute_steps(step)
                
                if context.break_loop:
                    logger.debug("Breaking step loop")
                    break
                
            # After step one iteration with loop change continue_loop = False
            context.continue_loop = False
            
            if context.break_loop:
                logger.debug("Breaking iteration loop")
                break
                
    except Exception as error:
        err = True
        err_msg = error

    context.break_loop = False
    context.continue_loop = False
    # update current step index to Loop again
    context.counters['index'] = currentStepIndex
    # set jumpLoop value to steps count
    context.jumpLoopIndex = len(steps)
    # remove 1 execution from executedStepsInLoop
    context.executedStepsInLoop -= context.jumpLoopIndex

    if err:
        raise CustomError(err_msg)

# Concludes the loop initiated with the step 'Loop "{x}" times starting at "{index}" and do'
# Example: Concludes the loop initiated with the step 'Loop "{x}" times starting at "{index}" and do'
@step(u'End Loop')
@done(u'End Loop')
def step_endLoop(context):
    try:
        # remove index variable from context.JOB_PARAMETERS
        params = json.loads(context.PARAMETERS)
        del params['index']
        context.PARAMETERS = json.dumps(params)
    except Exception as error:
        logger.error("Some error occured while trying to remove 'index' parameter from parameters. Maybe the loop had 0 iterations.")
        logger.debug("Here is the list of all parameters:")
        logger.debug(context.PARAMETERS)
    # set context.insideLoop to false
    context.insideLoop = False
    # reset jumpLoopIndex
    context.jumpLoopIndex = 0
    # reset executedStepsInLoop
    context.executedStepsInLoop = 0
    
# Example: Break the loop execution
@step(u'Break Loop')
@done(u'Break Loop')
def step_break_loop(context):
    send_step_details(context, "Breaking execution loop")
    
    if len(context.test_conditions_list)>0 and context.test_conditions_list[-1].is_condition_with_in_loop():            
        logger.debug("loop contains an if condition, closing it for this loop cycle")
        last_condition = context.test_conditions_list.pop()    
        last_condition.close_condition()
    
    if context.insideLoop:
        context.break_loop = True
    else:
        raise CustomError("Flows is not inside loop, invalid use of step")
    
    
# Example: Break the loop execution
@step(u'Continue Loop')
@done(u'Continue Loop')
def step_break_loop(context):
    send_step_details(context, "Continue this iteration of loop")
    
    if len(context.test_conditions_list)>0 and context.test_conditions_list[-1].is_condition_with_in_loop():            
        logger.debug("loop contains an if condition, closing it for this loop cycle")
        last_condition = context.test_conditions_list.pop()    
        last_condition.close_condition()
    
    if context.insideLoop:
        context.continue_loop = True
    else:
        raise CustomError("Flows is not inside loop, invalid use of step")
    


# This step tests if a list of elements selected via a CSS selector contains all or partial values derived from one or more variables
# Example: Test list of ".css_selector_example" elements to contain all or partial values from list variable "variable1|variable2" use prefix "prefix_" and suffix "_suffix"
@step(u'Test list of "{css_selector}" elements to contain all or partial values from list variable "{variable_names}" use prefix "{prefix}" and suffix "{suffix}"')
@done(u'Test list of "{css_selector}" elements to contain all or partial values from list variable "{variable_names}" use prefix "{prefix}" and suffix "{suffix}"')
def step_imp(context, css_selector, variable_names, prefix, suffix):
    # get all the values from css_selector
    elements = waitSelector(context, "css", css_selector)
    # get elements values
    element_values = [(element.get_attribute("innerText") or element.get_attribute("value")).strip() for element in elements]

    # get the variables from the context
    env_variables = json.loads(context.VARIABLES)
    # check if variable_name is in the env_variables
    index = [i for i,_ in enumerate(env_variables) if _['variable_name'] in variable_names.split("|")]

    if len(index) == 0:
        raise CustomError("No variable found with names: %s" % ", ".join(variable_names.split("|")))

    # create a variables that will contain all the concated values from variables
    values = []
    # loop over indexes found while looking for variables
    for i in index:
        # get variable value from the variables
        variable_value = env_variables[i]['variable_value']
        # split it from semi-colon (;)
        values.extend(variable_value.split(";"))

    # add prefix and suffix to the values
    values = [("%s%s%s" % (prefix, value.strip(), suffix)).strip() for value in values]

    # equalize the lenght of both lists
    values_eq = values[:len(element_values)] if len(values) > len(element_values) else values
    element_values_eq = element_values[:len(values)] if len(element_values) > len(values) else element_values

    # sorted lists
    element_values_sorted = sorted(element_values)
    values_sorted = sorted(values)

    # equalize the lenght of both sorted lists
    values_sorted_eq = values_sorted[:len(element_values_sorted)] if len(values_sorted) > len(element_values_sorted) else values_sorted
    element_sorted_values_eq = element_values_sorted[:len(values_sorted)] if len(element_values_sorted) > len(values_sorted) else element_values_sorted

    # save date and time an later format it
    dateTime = datetime.datetime.now()

    # save lists to file and link it to the step
    fileName = "list_comparison_%s.csv" % dateTime.strftime("%Y%m%d%H%M%S%f")
    filePath = "%s/%s" % (context.downloadDirectoryOutsideSelenium, fileName)

    fileContent = [
        ["Feature ID", int(context.feature_id)],
        ["Feature Result ID", int(os.environ['feature_result_id'])],
        ["Step number (starts from 1)", context.counters['index'] + 1],
        ["Selector", css_selector],
        ["Date & Time", dateTime.strftime("%Y-%m-%d %H:%M:%S")],
        ["Overall Status", "Passed" if (values_eq == element_values_eq) or (values_sorted_eq == element_sorted_values_eq) else 'Failed'],
        ["Option", "All/Partial (Not changeable for now)"],
        [""],
        ["Comparison Details Unsorted:"],
        ["Element List", "Variable List", "Status"]
    ]

    # print element list and variable list
    for i in range(0, len(values) if len(values) >= len(element_values) else len(element_values)):
        # get value for values_eq
        try:
            val = values[i]
        except:
            val = ''
        # get value for element_values_eq
        try:
            element_val = element_values[i]
        except:
            element_val = ''
        # get the result
        if val == element_val:
            result = "Same"
        else:
            result = "Not Same"

        if val == '' or element_val == '':
            result = "Ignored"

        fileContent.append([
            element_val, # Element List value
            val, # Variable List value,
            result # result of comparison
        ])

    # add additional text to fileContent
    fileContent.append([""])
    fileContent.append(["Comparison Details Sorted:"])
    fileContent.append(["Sorted Element List", "Sorted Variable List", "Status"])

    # print sorted element list and sorted variable list
    for i in range(0, len(values_sorted) if len(values_sorted) >= len(element_values_sorted) else len(element_values_sorted)):
        # get value for values_eq
        try:
            val = values_sorted[i]
        except:
            val = ''
        # get value for element_values_eq
        try:
            element_val = element_values_sorted[i]
        except:
            element_val = ''
        # get the result
        if val == element_val:
            result = "Same"
        else:
            result = "Not Same"

        if val == '' or element_val == '':
            result = "Ignored"

        fileContent.append([
            element_val, # Element List value
            val, # Variable List value,
            result # result of comparison
        ])

    # open file and write to it
    with open(filePath, 'w+', encoding="utf_8_sig") as fileHandle:
        writer = csv.writer(fileHandle, dialect="excel", delimiter=",", lineterminator='\n')
        writer.writerows(fileContent)

    # updated downloadedFiles in context
    context.downloadedFiles[context.counters['index']] = ["%s/%s" % (os.environ['feature_result_id'], fileName)]

    if (values_eq == element_values_eq) or (values_sorted_eq == element_sorted_values_eq):
        return True
    else:
        raise CustomError("Lists do not match, please check the attachment.")

# This step defines a custom error message to be applied to the next step in the scenario
# Example: Define Custom Error Message for next step: "The next step failed because of an unexpected error."
@step(u'Define Custom Error Message for next step: "{error_message}"')
@done(u'Define Custom Error Message for next step: "{error_message}"')
def step_imp(context, error_message):
    # get next step index
    next_step = context.counters['index'] + 1
    # check that there is another step after the current step
    logger.debug("Checking that there is another step")
    if len(context.steps) > next_step:
        # update the step definition
        context.steps[next_step].update({
            "custom_error": logger.mask_values(error_message)
        })
        logger.info(f"Custom error message set for step: {context.steps[next_step]['step_content']}")
    else:
        logger.warn(f"This is the last step, cannot assign custom error message to next step.")

# This step fetches the browser's console log and attaches it to the feature result
# Example: Fetch Console.log from Browser and attach it to the feature result
@step(u'Fetch Console.log from Browser and attach it to the feature result')
@done(u'Fetch Console.log from Browser and attach it to the feature result')
def attach_console_logs(context):
    if context.browser.capabilities.get('browserName', None) == 'firefox':
        notes = "Console logs are not reachable in firefox, please try another browser."
    else:
        logs = context.browser.get_log('browser')
        notes = ""
        for log in logs:
            date = datetime.datetime.utcfromtimestamp(log['timestamp'] / 1000) # divide it my 1000 since JS timezone in in ms.
            parse_log = re.search(r'^(?P<script_url>.*?) (?P<line_column>.*?) \"?(?P<message>.*?)\"?$', log['message'], flags=re.M | re.S)
            if not parse_log:
                continue
            message = parse_log.groupdict().get('message', '---')

            notes += f"[{date.strftime('%Y-%m-%d %H:%M:%S')}][{log['level']}] - {message}\n"

    context.step_data['notes'] = {
        'title': "Browser Console Logs",
        'content': notes
    }

# This step simulates dragging an element from one location and dropping it in a destination location on the page
# Example: Drag "element_selector_xpath" and drop it in "destination_selector_xpath"
@step(u'Drag "{element_selector}" and drop it in "{destination_selector}"')
@done(u'Drag "{element_selector}" and drop it in "{destination_selector}"')
def drag_n_drop(context, element_selector, destination_selector):
    element = waitSelector(context, "xpath", element_selector)
    destination = waitSelector(context, "xpath", destination_selector)

    if isinstance(element, list) and len(element) > 0:
        element = element[0]
    if isinstance(destination, list) and len(destination) > 0:
        destination = destination[0]

    ActionChains(context.browser).click_and_hold(element).move_to_element(destination).release(destination).perform()

# This step simulates selecting a checkbox if it's not already selected.
@step(u'Enable checkbox "{selector}"')
@done(u'Enable checkbox "{selector}"')
def select_checkbox(context, selector):
    checkbox = waitSelector(context, "xpath", selector)
    send_step_details(context, "Enabling checkbox")
    # Handle if returned as a list (some frameworks wrap in lists)
    if isinstance(checkbox, list) and len(checkbox) > 0:
        checkbox = checkbox[0]

    # Ensure checkbox is only clicked if not already selected
    if not checkbox.is_selected():
        checkbox.click()


@step(u'Disable checkbox "{selector}"')
@done(u'Disable checkbox "{selector}"')
def unselect_checkbox(context, selector):
    checkbox = waitSelector(context, "xpath", selector)
    send_step_details(context, "Disabling checkbox")
    if isinstance(checkbox, list) and len(checkbox) > 0:
        checkbox = checkbox[0]

    if checkbox.is_selected():
        checkbox.click()        

@step(u'Validate if checkbox "{selector}" is enabled save result in "{variable}"')
@done(u'Validate if checkbox "{selector}" is enabled save result in "{variable}"')
def unselect_checkbox(context, selector, variable):
    checkbox = waitSelector(context, "xpath", selector)
    send_step_details(context, "Validate if checkbox is enabled")
    if isinstance(checkbox, list) and len(checkbox) > 0:
        checkbox = checkbox[0]

    # Save the result to the variable
    addTestRuntimeVariable(context, variable, str(checkbox.is_selected()),  save_to_step_report=True)
        
        
# This step fetches the HTML source code of the current browser page and attaches it to the feature result as notes
# Example: Fetch HTML Source of current Browser page and attach it to the feature result
@step(u'Fetch HTML Source of current Browser page and attach it to the feature result')
@done(u'Fetch HTML Source of current Browser page and attach it to the feature result')
def fetch_page_source(context):
    context.step_data['notes'] = {
        'title': "Page Source Content",
        'content': context.browser.page_source
    }

# This step scrolls to a specific element in a lazy-loaded table (such as one in AG Grid)
# Scroll to element in lazy-loaded table, specially useful when working with AG Grid
@step(u'Scroll to element with "{selector}" in AG Grid table "{scrollable_element_selector}"')
@done(u'Scroll to element with "{selector}" in AG Grid table "{scrollable_element_selector}"')
def find_element_in_lazy_loaded_element(context, selector, scrollable_element_selector):

    # get the scrollable element
    scrollable_elements = waitSelector(context, 'xpath', scrollable_element_selector)
    if len(scrollable_elements) > 0:
        scrollable_element = scrollable_elements[0]
    else:
        raise CustomError("Unable to find scrollable element.")
    # get the height of the scrollable element
    total_scrollable_height = scrollable_element.rect['height']

    # get scrollable element parent node
    # scrollable_element_parent: WebElement = context.browser.execute_script('return arguments[0].parentNode', scrollable_element)
    # total height of the parent element
    # total_scrollable_parent_height = scrollable_element_parent.rect['height']

    # how many scrolls are required in total?
    # total_scrolls = int(total_scrollable_height/total_scrollable_parent_height)
    # scroll it by 10% - not a good solution...
    percentual_height = total_scrollable_height * 0.1
    total_scrolls = int(total_scrollable_height/percentual_height)

    # is element found or not
    found = False

    # scroll and find the element
    for i in range(0, total_scrolls):
        scroll_amount = int(i * percentual_height)
        scroll_origin = ScrollOrigin.from_element(scrollable_element)
        ActionChains(context.browser).scroll_from_origin(scroll_origin, 0, scroll_amount).perform()
        try:
            element = waitSelector(context, 'css', selector, 5)
            if isinstance(element, list) and len(element) > 0:
                element = element[0]
            if element:
                found = True
                # bring the element into the view
                ActionChains(context.browser).scroll_to_element(element).perform()
                break
        except CometaMaxTimeoutReachedException:
            pass

    if not found:
        raise CustomError('Element not found in lazy-loaded table... please try a different selector.')



# Read tag value from given selector and store in a variable
# Example 1: Get value from "//a[@href="/home"]" and store in "home_link_text" with ""
# Example 2: Get value from "//a[@href="/home"]" and store in "home_link_text" with "trim the spaces"
@step(u'Get value from "{selector}" and store in the "{variable_name}" with "{option}"')
@done(u'Get value from "{selector}" and store in the "{variable_name}" with "{option}"')
def step_impl(context, selector, variable_name, option):
    send_step_details(context, f'Searching for selector: {selector}')

    if option not in ['', 'trim the spaces']:
        raise ValueError(f'Invalid option "{option}". Use "", or "trim the spaces".')

    elements = waitSelector(context, "css", selector)

    if not elements:    
        raise ValueError(f"No elements found for selector: {selector}")

    try:
        # Get the tag value (text content)
        tag_value = elements[0].text
        logger.debug(f"Tag value found: {tag_value}")

        # Process the tag value based on the given option
        if option.lower() == "trim the spaces":
            tag_value = tag_value.strip()

        # Store the processed value in the variable
        addVariable(context, variable_name, tag_value, save_to_step_report=True)
        send_step_details(context, f'Stored value "{tag_value}" in variable "{variable_name}".')

    except Exception as err:
        raise CustomError(err)


# Read tag's attribute value from given selector and store in a variable
# Example 1: Get "href" value of "//a[@href="/home"]" and store in "href_value" with ""
# Example 2: Get "class" value of "//a[@href="/home"]" and store in "class_value" with "trim the spaces"
@step(u'Get "{attribute}" value of "{selector}" and store in the "{variable_name}" with "{option}"')
@done(u'Get "{attribute}" value of "{selector}" and store in the "{variable_name}" with "{option}"')
def step_impl(context, attribute, selector, variable_name, option):
    send_step_details(context, f'Searching for selector: {selector}')

    if option not in ['', 'trim the spaces']:
        raise ValueError(f'Invalid option "{option}". Use "", or "trim the spaces".')

    elements = waitSelector(context, "css", selector)

    if not elements:
        raise ValueError(f"No elements found for selector: {selector}")

    try:
        # Get the attribute value
        attribute_value = elements[0].get_attribute(attribute)

        if attribute_value is None:
            raise ValueError(f'Attribute "{attribute}" not found for selector: {selector}')

        # Process the attribute value based on the given option
        if option.lower() == "trim the spaces":
            attribute_value = attribute_value.strip()

        # Store the processed value in the variable
        addVariable(context, variable_name, attribute_value, save_to_step_report=True)
        send_step_details(context, f'Stored value "{attribute_value}" in variable "{variable_name}".')

    except Exception as err:
        raise CustomError(err)





# This step waits for the selector to appear within the given timeout (in seconds) and then wait the step timeout to disappear. The options provided are 'do not fail if not visible' or 'fail if never visible'
# If the selector does not appear within the specified timeout:
# 1. If the selected option is 'do not fail if not visible', the step will not fail, and it will skip the wait to disappear.
# 2. If the selected option is 'fail if never visible', the step will fail.
# 3. If the selected option is 'reload page after appearing', then the page is reloaded 0.5 seconds after the selector appear
# 4. If the selected option is 'reload page while waiting to disappear', then the page is reloaded every minute while waiting to disappear. 
# Options can be chained like:  'reload page after appearing;do not fail if not visible' 
# For Option 4: the step timeout must be bigger then 1 minute, as the waiting time after reloading is set to 1 minute.
@step(u'Wait "{timeout}" seconds for "{selector}" to appear and disappear using option "{option}"')
@done(u'Wait "{timeout}" seconds for "{selector}" to appear and disappear using option "{option}"')
def wait_for_appear_and_disappear(context, timeout, selector, option):

    assert_options = ['do not fail if not visible','fail if never visible']
    reload_options = ['reload page after appearing','reload page while waiting to disappear']

    options = option.split(";")

    # By default page reload in between is false
    page_reload = ''
    fail_option = ''

    # Read the options into static variables
    for option in options:
        option = option.strip()
        logger.debug(f"Found option: {option} - will evaluate if it is a valid option")
        if option in assert_options:
            fail_option = option
        elif option in reload_options :
            page_reload = option
        else:
            raise CustomError("Unknown option, options can be one of these options: %s." % ", ".join(reload_options+assert_options))


    # Make timeout a float
    timeout = float(timeout)

    # check if match type is one of reload_options
    if page_reload !='' and page_reload not in reload_options:
        raise CustomError("Unknown option, page_reload can be one of these options: %s." % ", ".join(reload_options))

    # Try ... except to handel option
    try:
        logger.debug("Waiting for selector to appear")
        send_step_details(context, 'Waiting for selector to appear')
        # Get the maximum wait timeout 
        max_time = time.time() + timeout
        selector_element = waitSelector(context, "css", selector, timeout)
        logger.debug("Got selector")

        # If element was loaded in dom but hidden, wait for it to display with max timeout
        while time.time()<max_time and len(selector_element)>0 and not selector_element[0].is_displayed():
            time.sleep(0.5)

        if len(selector_element)>0 and selector_element[0].is_displayed():
            send_step_details(context, 'Selector appeared, Wait for it to disappear')
            logger.debug("Waiting for selector to disappear")

            # Check if "reload page after appearing" options is set
            logger.debug("Checking for 'reload page after appearing'")
            # if option is set then sleep (0.5) and reload page
            if page_reload=='reload page after appearing':
                time.sleep(0.5)
                send_step_details(context, 'Reloading page after appearing')
                context.browser.refresh()
                logger.debug(f"Page reloaded - now waiting {timeout}s for selector to be visible")
                # This is required because when the page reloads, the selector_element does not remain valid.
                # It will throw a StaleElementReferenceException even if the element exists, try to get a new instance of the element.                        
                selector_element = waitSelector(context, "css", selector, timeout)

            # If in case object disappears and gets removed from DOM it self the while checking is_displayed() it will throw error
            # Considering error as element was disappeard
            try:
                # continue loop if element is displayed for wait time is less then 60 seconds
                while selector_element[0].is_displayed():
                    # Reload if "reload page while waiting to disappear" is set
                    if page_reload=='reload page while waiting to disappear':
                        send_step_details(context, 'Reloading the page while waiting for the selector to disappear')
                        logger.debug("Page reloading")
                        context.browser.refresh()
                        # If set - reload page and sleep 1 minute
                        send_step_details(context, 'Waiting for 1 minute after the page reloads')
                        logger.debug("Waiting for 1 minute after the page reloads")
                        time.sleep(60)
                        # This is required because when the page reloads, the selector_element does not remain valid.
                        # It will throw a StaleElementReferenceException even if the element exists, try to get a new instance of the element.
                        logger.debug(f"Now waiting 2s for selector to be visible")
                        # This throws CometaMaxTimeoutReachedException when selector does not appear in 2 seconds
                        selector_element = waitSelector(context, "css", selector, 2)
                    else:
                        time.sleep(0.5)
                # In case element disappeard but present in the DOM
                if not selector_element[0].is_displayed(): 
                    send_step_details(context, 'Selector disappeared successfully')
            except (StaleElementReferenceException, CometaMaxTimeoutReachedException) as e:
                # CometaMaxTimeoutReachedException exception handled here  
                # The only syntax that can throw error is is_displayed() method. which means element diappeared
                # In case element disappeard and not present in the DOM notify
                send_step_details(context, 'Selector disappeared completely')
                logger.debug("Element has disappeared due to exception")

        # if element was not found or displayed then check if option selected to fail, if yes then raise exception
        elif fail_option == 'fail if never visible':
            logger.debug("raising error: Selector to appeared")
            raise CustomError("Selector not displayed")

    except (StaleElementReferenceException, CometaMaxTimeoutReachedException) as exception:
        logger.error("Found error - StaleException or TimeoutReached")
        traceback.print_exc()
        # if got execption while checking to appear or disappear then check if option selected to fail, if yes then raise exception
        if fail_option == 'fail if never visible':
            logger.error("Raising error: fail if never visible")
            raise exception

# This step scrolls through a lazy-loaded table (or page) to reach the last visible position of a specified element, identified by its XPath
# Scroll to the end of the page/table depending on the xpath with maximum scrolls and time of life.
@step(u'Scroll to the last position of the desired element identified by "{xpath}" with maximum number of scrolls "{MaxScrolls}" and maximum time of  "{MaxTimeOfLife}"')
@done(u'Scroll to the last position of the desired element identified by "{xpath}" with maximum number of scrolls "{MaxScrolls}" and maximum time of  "{MaxTimeOfLife}"')
def scrollThroughLazyLoading(context, xpath, MaxScrolls, MaxTimeOfLife):
    context.browser.execute_script('''
    const XPathSelector = "%s"; // Define the XPath selector to catch the relative final argument of your table
    function wait(milliseconds) {
    return new Promise(resolve => {
        setTimeout(() => {resolve("")}, milliseconds);
    });
    }
    async function scrollToEnd(maxScrolls, maxTimeInSeconds) {
    let lastScrollPosition = -1;
    let tries = 0;
    let numberOfScrolls = 0;
    const startTime = Date.now();
    const maximumScrolls = Math.abs(maxScrolls);
    const maxTime = Math.abs(maxTimeInSeconds) * 1000;
    async function doesScroll() {
        if (numberOfScrolls >= maximumScrolls || (Date.now() - startTime) > maxTime) {
            return;
        }
        let lastElement = document.evaluate(XPathSelector, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;
        if (lastElement) {
            lastElement.scrollIntoView();
        }

        if (window.scrollY === lastScrollPosition) {
            tries++;
            if (tries > 4) { // Sets max amount of scrolling attempts before exiting
            return;
            }
        } else {
            lastScrollPosition = window.scrollY;
            tries = 0;
            numberOfScrolls++;
        }
        await wait(500); // Wait time between scrolling attempts in miliseconds
        await doesScroll();
    }
    await doesScroll();
    }
    const MaxNumberOfScrolls = %s; // Define max number of total scrolls
    const MaxTimeOfWaiting = %s;   // Define JS script time of life in seconds
    await scrollToEnd(MaxNumberOfScrolls, MaxTimeOfWaiting);
    ''' % (xpath, MaxScrolls, MaxTimeOfLife))

# This step reads the html source code of the current page, and looks for aria label attributes, it counts the number of occurrences of each aria label attribute, and displays the value in the options section of the step report.
# Count the number of aria labels in the current page
@step(u'Count the number of aria labels in the current page')
@done(u'Count the number of aria labels in the current page')
def count_aria_labels(context):
    # Get the html source code of the current page
    html_source = context.browser.page_source
    # Count the number of aria labels in the html source code
    aria_labels = re.findall(r'aria-label="([^"]+)"', html_source)
    # Count the number of occurrences of each aria label
    aria_label_counts = {}
    for label in aria_labels:
        aria_label_counts[label] = aria_label_counts.get(label, 0) + 1
    # Display the value in the options section of the step report
    context.step_data['notes'] = {
        'title': "Aria Labels",
        'content': json.dumps(aria_label_counts)  # Convert the dictionary to a JSON string
    }



@step(u'Enable browser network logging')
@done(u'Enable browser network logging')
def fetch_all_network_requests(context):
      # Enable Network tracking
    context.browser.execute(
        "executeCdpCommand",
        {
            "cmd": "Network.enable",
            "params": {}
        }
    )

@step(u'Disable browser network logging')
@done(u'Disable browser network logging')
def fetch_all_network_requests(context):
      # Enable Network tracking
    context.browser.execute(
        "executeCdpCommand",
        {
            "cmd": "Network.disable",
            "params": {}
        }
    )
    

@step(u'Get all browser network requests and save in variable "{variable}"')
@done(u'Get all browser network requests and save in variable "{variable}"')
def fetch_all_network_requests(context, variable):
    # Dictionaries to store request and response info by requestId
    requests = {}
    responses = {}
    logger.debug("fetching and processing network logs")

    logs = context.browser.get_log('performance')
    for entry in logs:
        log = json.loads(entry['message'])['message']
        method = log.get('method')
        params = log.get('params', {})

        # Collect request info
        if method == 'Network.requestWillBeSent':
            request_id = params['requestId']
            request = params['request']
            requests[request_id] = request
                

        # Collect response info
        elif method == 'Network.responseReceived':
            request_id = params['requestId']
            response = params['response']
            responses[request_id] = response

    # For each request that has a response, get the response body
    for request_id in responses:
        # Get response body via CDP
        try:
            response_body = context.browser.execute(
                "executeCdpCommand",
                {
                    "cmd": "Network.getResponseBody",
                    "params": {"requestId": request_id}
                }
            )
            responses[request_id]['body'] = response_body.get('value', {}).get('body')
        except Exception:
            responses[request_id]['body'] = None

    all_requests_data = []
    logger.debug("Combining the request and response")
    # Combine and store info for all requests
    for request_id in requests:
        req = requests[request_id]
        resp = responses.get(request_id)
        entry = {
            "request_id": request_id,
            "url": req['url'],
            "request_headers": req['headers'],
            "request_body": req.get('body'),
            "request_method": req.get('method'),
        }
        if resp:
            entry.update({
                "response_status": resp.get('status'),
                "response_headers": resp.get('headers'),
                "response_body": resp.get('body'),
            })
        else:
            entry.update({
                "response_status": None,
                "response_headers": None,
                "response_body": None,
            })
        all_requests_data.append(entry)
    logger.debug(f"Total {len(all_requests_data)} requests found")
    addTestRuntimeVariable(context, variable, all_requests_data, save_to_step_report=True)



@step(u'Get browser network requests filter by "{url}" and save in the variable "{variable}"')
@done(u'Get browser network requests filter by "{url}" and save in the variable "{variable}"')
def fetch_network_requests_by_url(context, url,  variable):
    # Dictionaries to store request and response info by requestId
    logger.debug("fetching and processing network logs")
    requests = {}
    responses = {}

    # Temporary map to link requestId to matching URLs
    matching_request_ids = set()

    logs = context.browser.get_log('performance')
    for entry in logs:
        log = json.loads(entry['message'])['message']
        method = log.get('method')
        params = log.get('params', {})

        # Collect request info, filter by URL
        if method == 'Network.requestWillBeSent':
            request_id = params['requestId']
            request = params['request']
            if url in request['url']:
                requests[request_id] = request
                matching_request_ids.add(request_id)

        # Collect response info only for matching requests
        elif method == 'Network.responseReceived':
            request_id = params['requestId']
            if request_id in matching_request_ids:
                response = params['response']
                responses[request_id] = response

    # For each request that has a response, get the response body
    for request_id in responses:
        try:
            response_body = context.browser.execute(
                "executeCdpCommand",
                {
                    "cmd": "Network.getResponseBody",
                    "params": {"requestId": request_id}
                }
            )
            responses[request_id]['body'] = response_body.get('value', {}).get('body')
        except Exception:
            responses[request_id]['body'] = None

    all_requests_data = []
    logger.debug("Combining the request and response")
    # Combine and store info for all matching requests
    for request_id in requests:
        req = requests[request_id]
        resp = responses.get(request_id)
        entry = {
            "request_id": request_id,
            "url": req['url'],
            "request_headers": req['headers'],
            "request_body": req.get('body'),
            "request_method": req.get('method'),
        }
        if resp:
            entry.update({
                "response_status": resp.get('status'),
                "response_headers": resp.get('headers'),
                "response_body": resp.get('body'),
            })
        else:
            entry.update({
                "response_status": None,
                "response_headers": None,
                "response_body": None,
            })
        all_requests_data.append(entry)
    
    logger.debug(f"Total {len(all_requests_data)} requests found")
    addTestRuntimeVariable(context, variable, all_requests_data, save_to_step_report=True)


# This step performs a basic accessibility check for aria-labels on the current page, providing a report on compliance with WCAG guidelines.
# The report includes counts of elements with and without aria-labels, recommendations for improvement, and a basic compliance score.
# Example: Check accessibility compliance for aria-labels on current page
@step(u'Check accessibility compliance for aria-labels on current page')
@done(u'Check accessibility compliance for aria-labels on current page')
def check_aria_labels_accessibility(context):
    # Get the html source code of the current page
    html_source = context.browser.page_source
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html_source, 'html.parser')
    import datetime
    import html

    # Helper: does this element have an accessible name?
    def has_accessible_name(elem):
        if elem.has_attr('aria-label') and elem['aria-label'].strip():
            return True
        if elem.has_attr('aria-labelledby') and elem['aria-labelledby'].strip():
            return True
        if elem.name == 'img' and elem.has_attr('alt') and elem['alt'].strip():
            return True
        if elem.get_text(strip=True):
            return True
        return False

    # 1. Only check non-semantic interactive elements (div, span, svg with role or onclick)
    interactive_candidates = soup.find_all(
        lambda tag: (
            tag.name in ['div', 'span', 'svg']
            and (tag.has_attr('onclick') or tag.has_attr('role'))
        )
    )

    # 2. Find those missing accessible names
    missing_aria_labels = []
    for elem in interactive_candidates:
        if not has_accessible_name(elem):
            elem_str = str(elem)[:100] + ("..." if len(str(elem)) > 100 else "")
            missing_aria_labels.append({
                'element_type': elem.name,
                'element': elem_str
            })

    # 3. Count all aria-labels (for reporting and linter)
    elements_with_aria_label = soup.find_all(attrs={"aria-label": True})
    aria_label_values = [elem.get('aria-label', '').strip() for elem in elements_with_aria_label if elem.get('aria-label', '').strip()]
    aria_label_counts = {}
    for label in aria_label_values:
        aria_label_counts[label] = aria_label_counts.get(label, 0) + 1

    # 4. Lint for non-descriptive labels
    common_bad_labels = {'click here', 'button', 'link'}
    bad_label_warnings = []
    for label in aria_label_counts:
        if label.lower() in common_bad_labels:
            bad_label_warnings.append(f'Consider replacing generic aria-label "{label}" with a more descriptive label')

    # 5. Report stats
    total_checked = len(interactive_candidates)
    total_missing = len(missing_aria_labels)
    compliance_score = round(((total_checked - total_missing) / total_checked) * 100, 2) if total_checked > 0 else 100
    if compliance_score >= 70:
        compliance_status = "Good"
        status_color = "green"
    elif compliance_score >= 50:
        compliance_status = "Average"
        status_color = "orange"
    else:
        compliance_status = "Poor"
        status_color = "red"

    # 6. Recommendations
    recommendations = []
    if missing_aria_labels:
        recommendations.append("Add aria-label, aria-labelledby, or visible text to interactive non-semantic elements (div/span/svg with role/onclick)")
    if bad_label_warnings:
        recommendations.extend(bad_label_warnings)
    if len(aria_label_counts) < len(elements_with_aria_label):
        recommendations.append("Avoid duplicate aria-labels for different functionality")

    # 7. HTML report
    html_report = f"""
    <html>
    <head>
        <title>Accessibility Compliance Report: ARIA Labels</title>
    </head>
    <body>
        <h1>Accessibility Compliance Report: ARIA Labels</h1>
        <div class="scorecard">
            <div class="score-circle">
                <h2 class="{status_color}">{compliance_score}%</h2>
                <div class="{status_color}">{compliance_status}</div>
            </div>
            <div class="stats">
                <h3>Non-semantic Interactive Elements Checked</h3>
                <p>Page tested: {context.browser.current_url}</p>
                <p>Date: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
            </div>
        </div>
        <h2>Summary</h2>
        <table>
            <tr>
                <th>Total Interactive Non-Semantic Elements</th>
                <th>With Accessible Name</th>
                <th>Without Accessible Name</th>
                <th>Compliance Score</th>
            </tr>
            <tr>
                <td>{total_checked}</td>
                <td>{total_checked - total_missing}</td>
                <td>{total_missing}</td>
                <td class="{status_color}">{compliance_score}% ({compliance_status})</td>
            </tr>
        </table>
        <h2>ARIA Label Usage</h2>
        <table>
            <tr>
                <th>ARIA Label Text</th>
                <th>Occurrences</th>
            </tr>
    """
    for label, count in aria_label_counts.items():
        html_report += f"""
            <tr>
                <td>{label}</td>
                <td>{count}</td>
            </tr>
        """
    html_report += """
        </table>
    """
    if missing_aria_labels:
        from collections import Counter

        # Count element occurrences
        element_counter = Counter(m['element'] for m in missing_aria_labels)

        # Build a new list with annotated entries
        annotated_missing = []
        for m in missing_aria_labels:
            elem_text = m['element']
            count = element_counter[elem_text]
            if count > 1:
                annotated_elem = f"{elem_text} (repeated {count} times)"
            else:
                annotated_elem = elem_text

            # Handle truly empty string (edge case)
            if not annotated_elem.strip():
                annotated_elem = "[Empty Element]"

            annotated_missing.append({
                'element_type': m['element_type'],
                'element': annotated_elem
            })

        html_report += """
        <h2>Missing Accessible Names (Examples)</h2>
        <table>
            <tr>
                <th>Element Type</th>
                <th>Element</th>
            </tr>
        """

        for i, missing in enumerate(annotated_missing[:10]):
            html_report += f"""
                <tr>
                    <td>{missing['element_type']}</td>
                    <td><code>{html.escape(missing['element'])}</code></td>
                </tr>
            """

        if len(annotated_missing) > 10:
            html_report += f"""
                <tr>
                    <td colspan="2">And {len(annotated_missing) - 10} more...</td>
                </tr>
            """

        html_report += """
        </table>
        """

    if recommendations:
        html_report += """
        <h2>Recommendations</h2>
        <ul>
        """
        for recommendation in recommendations:
            html_report += f"""
            <li>{recommendation}</li>
            """
        html_report += """
        </ul>
        """
    html_report += """
        <h2>Relevant WCAG Guidelines</h2>
        <ul>
            <li>WCAG 2.1 Success Criterion 1.1.1: Non-text Content (Level A)</li>
            <li>WCAG 2.1 Success Criterion 2.4.4: Link Purpose (In Context) (Level A)</li>
            <li>WCAG 2.1 Success Criterion 4.1.2: Name, Role, Value (Level A)</li>
        </ul>
    </body>
    </html>
    """
    context.step_data['notes'] = {
        'title': "Accessibility Compliance Report: ARIA Labels",
        'content': html_report
    }

if __name__ != 'actions':
    sys.path.append('/opt/code/')
    from ee.cometa_itself.steps import rest_api  
    from ee.cometa_itself.steps import conditional_actions
    COMETA_FEATURE_DATABASE_ENABLED = ConfigurationManager.get_configuration("COMETA_FEATURE_MOBILE_TEST_ENABLED","False")=="True"
    # logger.debug(f"COMETA_FEATURE_DATABASE_ENABLED : {COMETA_FEATURE_DATABASE_ENABLED}")
    if COMETA_FEATURE_DATABASE_ENABLED:
        from ee.cometa_itself.steps import mobile_actions 
        logger.debug("Importing mobile_actions")
        
    COMETA_FEATURE_DATABASE_ENABLED = ConfigurationManager.get_configuration("COMETA_FEATURE_DATABASE_ENABLED","False")=="True"
    # logger.debug(f"COMETA_FEATURE_DATABASE_ENABLED : {COMETA_FEATURE_DATABASE_ENABLED}")
    if COMETA_FEATURE_DATABASE_ENABLED:
        logger.debug("Importing database_actions")
        from ee.cometa_itself.steps import database_actions  
    
    COMETA_FEATURE_AI_ENABLED = ConfigurationManager.get_configuration("COMETA_FEATURE_AI_ENABLED","False")=="True"
    # logger.debug(f"COMETA_FEATURE_AI_ENABLED : {COMETA_FEATURE_AI_ENABLED}")
    if COMETA_FEATURE_AI_ENABLED:
        from ee.cometa_itself.steps import ai_actions
        from ee.cometa_itself.steps import playwright_actions
        logger.debug("Importing ai_actions")        


    from ee.cometa_itself.steps import common_actions  
    
    sys.path.append('/opt/code/cometa_itself')
    from steps import validation_actions
    from steps import unimplemented_steps
    
